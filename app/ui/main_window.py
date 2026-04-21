from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from datetime import date, timedelta
from pathlib import Path

from PySide6.QtCore import QDate, QSize, Qt, QTimer
from PySide6.QtGui import QColor, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QFrame,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QSpinBox,
    QPushButton,
    QSizePolicy,
    QScrollArea,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from app.services.bookkeeping_service import BookkeepingService
from app.services.inventory_service import InventoryService
from app.services.print_service import PrintService
from app.services.report_service import ReportService
from app.services.sales_service import SalesService
from app.utils.backup import create_backup, export_backup, inspect_backup_counts, restore_backup


class MainWindow(QMainWindow):
    MIN_WINDOW_WIDTH = 900
    MIN_WINDOW_HEIGHT = 680

    def __init__(
        self,
        inventory_service: InventoryService,
        sales_service: SalesService,
        bookkeeping_service: BookkeepingService,
        report_service: ReportService,
        print_service: PrintService,
        db_path: str = "data/cafe.db",
    ) -> None:
        super().__init__()
        self.inventory_service = inventory_service
        self.sales_service = sales_service
        self.bookkeeping_service = bookkeeping_service
        self.report_service = report_service
        self.print_service = print_service
        self.db_path = db_path

        self.cart: dict[int, dict] = {}
        self.billing_items_cache: list[dict] = []
        self.inventory_items_cache: list[dict] = []
        self.cigarette_shortcuts: list[QShortcut] = []
        self.purchase_cart: list[dict] = []
        self.purchase_item_cache: dict[int, dict] = {}
        self.editing_purchase_id: int | None = None
        self.current_role = (self.bookkeeping_service.get_setting("current_role", "cashier") or "cashier").lower()
        self._updating_inventory_table = False
        self._updating_purchase_table = False
        self._updating_expense_table = False
        self.cart_file = Path("data/pending_cart.json")

        self.setWindowTitle("Cafe POS and Bookkeeping")
        self.setMinimumSize(self.MIN_WINDOW_WIDTH, self.MIN_WINDOW_HEIGHT)
        self.resize(1320, 820)

        self._build_professional_shell_header()

        self.tabs = QTabWidget()
        self.tabs.setObjectName("AppTabs")
        self.setCentralWidget(self.tabs)

        self.billing_tab = self._build_billing_tab()
        self.inventory_tab = self._build_inventory_tab()
        self.purchases_tab = self._build_purchases_tab()
        self.expenses_tab = self._build_expenses_tab()
        self.reports_tab = self._build_reports_tab()

        self.tabs.addTab(self.billing_tab, "Billing")
        self.tabs.addTab(self.inventory_tab, "Inventory")
        self.tabs.addTab(self.purchases_tab, "Purchases")
        self.tabs.addTab(self.expenses_tab, "Expenses")
        self.tabs.addTab(self.reports_tab, "Reports")
        self.tabs.currentChanged.connect(self._on_tab_changed)

        self._apply_professional_shell_theme()
        self._setup_status_bar()

        self._wire_shortcuts()

        self.cart_timer = QTimer(self)
        self.cart_timer.setInterval(5000)
        self.cart_timer.timeout.connect(self._save_pending_cart)
        self.cart_timer.start()

        self.auto_backup_timer = QTimer(self)
        self.auto_backup_timer.timeout.connect(self._run_scheduled_backup)
        self._configure_auto_backup_timer()

        self.shell_clock_timer = QTimer(self)
        self.shell_clock_timer.setInterval(1000)
        self.shell_clock_timer.timeout.connect(self._update_shell_status)
        self.shell_clock_timer.start()

        self.refresh_all()
        self._load_pending_cart()
        self._update_shell_status()

    def minimumSizeHint(self) -> QSize:
        # Keep window constraints practical so the app can fit smaller screens.
        return QSize(self.MIN_WINDOW_WIDTH, self.MIN_WINDOW_HEIGHT)

    def showEvent(self, event) -> None:
        super().showEvent(event)
        QTimer.singleShot(0, self._fit_window_to_available_screen)

    def _fit_window_to_available_screen(self) -> None:
        screen = self.screen() or QApplication.primaryScreen()
        if screen is None:
            return

        available = screen.availableGeometry()
        max_width = max(640, available.width())
        max_height = max(520, available.height())

        # Never exceed available geometry so taskbar does not overlay content.
        self.setMaximumSize(max_width, max_height)

        min_width = min(self.MIN_WINDOW_WIDTH, max_width)
        min_height = min(self.MIN_WINDOW_HEIGHT, max_height)
        self.setMinimumSize(min_width, min_height)

        target_width = min(max(self.width(), min_width), max_width)
        target_height = min(max(self.height(), min_height), max_height)
        self.setGeometry(available.x(), available.y(), target_width, target_height)

    def _on_tab_changed(self, tab_index: int) -> None:
        if tab_index == 0:
            self.refresh_billing_items()
            QTimer.singleShot(0, self._focus_billing_search)
        elif tab_index == 1:
            self.refresh_inventory()
            QTimer.singleShot(0, self._focus_inventory_name)
        elif tab_index == 2:
            self.refresh_purchases_tab()
        elif tab_index == 3:
            self.refresh_expenses_tab()
        elif tab_index == 4:
            self.refresh_reports()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        if hasattr(self, "top_items_table"):
            self._apply_report_table_width_profiles()
        if hasattr(self, "billing_compact_toggle") and self.billing_compact_toggle.isChecked():
            self._set_billing_compact_mode(True)

    def _wire_shortcuts(self) -> None:
        QShortcut(QKeySequence("Ctrl+Return"), self, activated=self.checkout)
        QShortcut(QKeySequence("F5"), self, activated=self.refresh_all)
        QShortcut(QKeySequence("Ctrl+L"), self, activated=self.close_day)
        QShortcut(QKeySequence("Ctrl+F"), self, activated=self._focus_billing_search)
        QShortcut(QKeySequence("Ctrl+Q"), self, activated=self._focus_billing_qty)
        QShortcut(QKeySequence("Return"), self, activated=self._try_checkout_from_enter)
        QShortcut(QKeySequence("F11"), self, activated=self._toggle_full_screen)

    def _toggle_full_screen(self) -> None:
        if self.isFullScreen():
            self.showMaximized()
            return
        self.showFullScreen()

    def _build_professional_shell_header(self) -> None:
        header = QWidget()
        header.setObjectName("AppShellHeader")
        layout = QHBoxLayout(header)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(10)

        title = QLabel("Cafe POS")
        title.setObjectName("AppShellTitle")
        subtitle = QLabel("Retail Operations Console")
        subtitle.setObjectName("AppShellSubtitle")

        title_stack = QVBoxLayout()
        title_stack.setContentsMargins(0, 0, 0, 0)
        title_stack.setSpacing(1)
        title_stack.addWidget(title)
        title_stack.addWidget(subtitle)

        self.header_role_badge = QLabel("ROLE: CASHIER")
        self.header_role_badge.setObjectName("HeaderRoleBadge")

        layout.addLayout(title_stack)
        layout.addStretch()
        layout.addWidget(self.header_role_badge)

        self.setMenuWidget(header)

    def _apply_professional_shell_theme(self) -> None:
        self.setStyleSheet(
            "QMainWindow {"
            "background-color: #20242c;"
            "}"
            "#AppShellHeader {"
            "background-color: #182233;"
            "border-bottom: 1px solid #2a3b57;"
            "}"
            "#AppShellTitle {"
            "font-size: 15px;"
            "font-weight: 700;"
            "color: #f2f6fc;"
            "}"
            "#AppShellSubtitle {"
            "font-size: 11px;"
            "color: #a8bad7;"
            "}"
            "#HeaderRoleBadge {"
            "padding: 4px 10px;"
            "border: 1px solid #4b5d7a;"
            "border-radius: 12px;"
            "background-color: #25344d;"
            "color: #dce8fb;"
            "font-weight: 700;"
            "}"
            "QTabWidget#AppTabs::pane {"
            "border-top: 1px solid #33435b;"
            "top: -1px;"
            "}"
            "QTabBar::tab {"
            "background-color: #232c3b;"
            "color: #c8d7ee;"
            "padding: 8px 14px;"
            "margin-right: 2px;"
            "border-top-left-radius: 6px;"
            "border-top-right-radius: 6px;"
            "border: 1px solid #39475f;"
            "}"
            "QTabBar::tab:selected {"
            "background-color: #2c3e5b;"
            "color: #ffffff;"
            "border-color: #506792;"
            "font-weight: 700;"
            "}"
            "QTabBar::tab:hover {"
            "background-color: #2b394f;"
            "}"
            "QStatusBar {"
            "background-color: #17202f;"
            "border-top: 1px solid #2d3b53;"
            "color: #c7d5ec;"
            "}"
            "QStatusBar QLabel {"
            "color: #c7d5ec;"
            "padding: 0 6px;"
            "}"
        )

    def _setup_status_bar(self) -> None:
        self.status_role_label = QLabel()
        self.status_tab_label = QLabel()
        self.status_db_label = QLabel(f"DB: {Path(self.db_path).name}")
        self.status_time_label = QLabel()

        self.statusBar().addPermanentWidget(self.status_role_label)
        self.statusBar().addPermanentWidget(self.status_tab_label)
        self.statusBar().addPermanentWidget(self.status_db_label)
        self.statusBar().addPermanentWidget(self.status_time_label)

    def _update_shell_status(self) -> None:
        role = self.current_role.upper() if self.current_role else "CASHIER"
        self.status_role_label.setText(f"ROLE: {role}")
        if hasattr(self, "header_role_badge"):
            self.header_role_badge.setText(f"ROLE: {role}")

        if hasattr(self, "tabs") and self.tabs.count() > 0:
            tab_name = self.tabs.tabText(self.tabs.currentIndex())
            self.status_tab_label.setText(f"MODULE: {tab_name}")

        self.status_time_label.setText(datetime.now().strftime("%d %b %Y  %I:%M:%S %p"))

    def _build_placeholder_tab(self, message: str) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        label = QLabel(message)
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        return tab

    def _build_billing_tab(self) -> QWidget:
        tab = QWidget()
        tab.setObjectName("BillingDashboard")
        root_layout = QVBoxLayout(tab)
        root_layout.setContentsMargins(10, 10, 10, 10)
        root_layout.setSpacing(8)

        hero_frame = QFrame()
        hero_frame.setObjectName("BillingHero")
        hero_layout = QHBoxLayout(hero_frame)
        hero_layout.setContentsMargins(12, 8, 12, 8)

        hero_text_layout = QVBoxLayout()
        hero_title = QLabel("Billing Dashboard")
        hero_title.setObjectName("BillingHeroTitle")
        hero_text_layout.addWidget(hero_title)

        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(6)
        lines_card, self.billing_cart_lines_value = self._make_billing_stat_card("LINES")
        qty_card, self.billing_cart_qty_value = self._make_billing_stat_card("UNITS")
        total_card, self.billing_total_value = self._make_billing_stat_card("BILL")
        stats_layout.addWidget(lines_card)
        stats_layout.addWidget(qty_card)
        stats_layout.addWidget(total_card)

        hero_layout.addLayout(hero_text_layout)
        hero_layout.addStretch()
        hero_layout.addLayout(stats_layout)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)

        catalog_group = QGroupBox("Item Catalog")
        catalog_layout = QVBoxLayout(catalog_group)
        catalog_layout.setSpacing(8)

        search_row = QHBoxLayout()
        search_row.setSpacing(8)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search item and press Enter...")
        self.search_input.setMinimumHeight(36)
        self.search_input.textChanged.connect(self.apply_billing_filter)
        self.search_input.returnPressed.connect(self.add_selected_item_to_cart)
        search_row.addWidget(QLabel("<b>Search</b>"))
        search_row.addWidget(self.search_input)
        catalog_layout.addLayout(search_row)

        self.billing_empty_state_label = QLabel("")
        self.billing_empty_state_label.setObjectName("BillingEmptyState")
        self.billing_empty_state_label.setVisible(False)
        catalog_layout.addWidget(self.billing_empty_state_label)

        self.billing_items_table = QTableWidget(0, 3)
        self.billing_items_table.setHorizontalHeaderLabels(["Name", "Price", "Stock"])
        self.billing_items_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.billing_items_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.billing_items_table.horizontalHeader().setStretchLastSection(True)
        self.billing_items_table.cellDoubleClicked.connect(self._on_catalog_double_click)
        self.billing_items_table.cellClicked.connect(self._on_catalog_single_click)
        self._style_billing_table(self.billing_items_table)

        quick_group = QGroupBox("Quick Add")
        self.billing_quick_group = quick_group
        quick_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
        quick_group.setMaximumHeight(270)
        quick_layout = QVBoxLayout(quick_group)
        quick_layout.setSpacing(6)

        self.small_buttons_layout = QHBoxLayout()
        self.medium_buttons_layout = QHBoxLayout()
        self.big_buttons_layout = QHBoxLayout()
        self.small_buttons_layout.setSpacing(8)
        self.medium_buttons_layout.setSpacing(8)
        self.big_buttons_layout.setSpacing(8)

        quick_layout.addWidget(QLabel("SMALL"))
        quick_layout.addLayout(self.small_buttons_layout)
        quick_layout.addWidget(QLabel("MEDIUM"))
        quick_layout.addLayout(self.medium_buttons_layout)
        quick_layout.addWidget(QLabel("BIG"))
        quick_layout.addLayout(self.big_buttons_layout)

        catalog_layout.addWidget(quick_group, 0)
        catalog_layout.addWidget(self.billing_items_table, 1)

        controls = QHBoxLayout()
        self.qty_spin = QDoubleSpinBox()
        self.qty_spin.setDecimals(2)
        self.qty_spin.setMinimum(0.01)
        self.qty_spin.setMaximum(1000)
        self.qty_spin.setValue(1)
        self.qty_spin.setMinimumHeight(34)

        add_btn = QPushButton("Add Selected")
        add_btn.setObjectName("SecondaryBillingButton")
        add_btn.clicked.connect(self.add_selected_item_to_cart)

        refresh_btn = QPushButton("Refresh Items")
        refresh_btn.setObjectName("SecondaryBillingButton")
        refresh_btn.clicked.connect(self.refresh_billing_items)

        self.billing_compact_toggle = QCheckBox("Compact for Small Screens")
        self.billing_compact_toggle.toggled.connect(self._set_billing_compact_mode)

        controls.addWidget(QLabel("<b>Qty</b>"))
        controls.addWidget(self.qty_spin)
        controls.addWidget(add_btn)
        controls.addWidget(refresh_btn)
        controls.addStretch()
        controls.addWidget(self.billing_compact_toggle)
        catalog_layout.addLayout(controls)
        catalog_layout.setStretch(3, 1)

        cart_group = QGroupBox("Current Bill")
        cart_layout = QVBoxLayout(cart_group)
        cart_layout.setSpacing(10)

        self.cart_table = QTableWidget(0, 5)
        self.cart_table.setHorizontalHeaderLabels(["Item ID", "Name", "Qty", "Unit", "Line Total"])
        self.cart_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.cart_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.cart_table.horizontalHeader().setStretchLastSection(True)
        self._style_billing_table(self.cart_table)
        self.cart_table.setMinimumHeight(360)

        cart_actions_row = QHBoxLayout()
        plus_qty_btn = QPushButton("+ Qty")
        plus_qty_btn.setObjectName("SecondaryBillingButton")
        plus_qty_btn.clicked.connect(self.increase_selected_cart_item_qty)

        minus_qty_btn = QPushButton("- Qty")
        minus_qty_btn.setObjectName("SecondaryBillingButton")
        minus_qty_btn.clicked.connect(self.decrease_selected_cart_item_qty)

        remove_selected_btn = QPushButton("Remove Selected")
        remove_selected_btn.setObjectName("SecondaryBillingButton")
        remove_selected_btn.clicked.connect(self.remove_selected_cart_item)

        cart_actions_row.addWidget(plus_qty_btn)
        cart_actions_row.addWidget(minus_qty_btn)
        cart_actions_row.addWidget(remove_selected_btn)
        cart_actions_row.addStretch()

        total_panel = QFrame()
        total_panel.setObjectName("BillingTotalPanel")
        total_panel_layout = QVBoxLayout(total_panel)
        total_panel_layout.setContentsMargins(10, 10, 10, 10)
        total_panel_layout.setSpacing(8)

        self.total_label = QLabel("TOTAL: INR 0.00")
        self.total_label.setObjectName("BillingTotalLabel")
        self.total_label.setAlignment(Qt.AlignCenter)

        clear_btn = QPushButton("Clear Cart")
        clear_btn.setObjectName("SecondaryBillingButton")
        clear_btn.clicked.connect(self.clear_cart)

        checkout_btn = QPushButton("Generate Bill (Ctrl+Enter)")
        checkout_btn.setObjectName("PrimaryBillingButton")
        checkout_btn.setMinimumHeight(44)
        checkout_btn.clicked.connect(self.checkout)

        totals_row = QHBoxLayout()
        totals_row.addWidget(self.total_label)
        actions_row = QHBoxLayout()
        actions_row.addWidget(clear_btn)
        actions_row.addWidget(checkout_btn)

        total_panel_layout.addLayout(totals_row)
        total_panel_layout.addLayout(actions_row)

        cart_layout.addWidget(self.cart_table)
        cart_layout.addLayout(cart_actions_row)
        cart_layout.addWidget(total_panel)

        splitter.addWidget(catalog_group)
        splitter.addWidget(cart_group)
        splitter.setSizes([520, 780])
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 6)

        root_layout.addWidget(hero_frame)
        root_layout.addWidget(splitter)
        self.billing_root_layout = root_layout
        self.billing_splitter = splitter
        self._apply_billing_dashboard_style(tab)
        self._set_billing_compact_mode(self.width() < 1180)
        self._update_billing_dashboard_metrics(total_amount=0.0)
        return tab

    def _focus_billing_search(self) -> None:
        self.tabs.setCurrentIndex(0)
        if hasattr(self, "search_input"):
            self.search_input.setFocus()
            self.search_input.selectAll()

    def _focus_inventory_name(self) -> None:
        if hasattr(self, "item_name_input"):
            self.item_name_input.setFocus()
            self.item_name_input.selectAll()

    def _focus_billing_qty(self) -> None:
        self.tabs.setCurrentIndex(0)
        if hasattr(self, "qty_spin"):
            self.qty_spin.setFocus()
            self.qty_spin.selectAll()

    def _try_checkout_from_enter(self) -> None:
        if self.tabs.currentIndex() != 0:
            return

        focused = self.focusWidget()
        if focused is self.search_input:
            return
        if focused is self.qty_spin or focused is self.qty_spin.lineEdit():
            return
        self.checkout()

    def _set_billing_compact_mode(self, enabled: bool) -> None:
        if not hasattr(self, "billing_root_layout"):
            return

        compact = bool(enabled)
        if compact:
            self.billing_root_layout.setContentsMargins(6, 6, 6, 6)
        else:
            self.billing_root_layout.setContentsMargins(10, 10, 10, 10)
        self.billing_root_layout.setSpacing(6 if compact else 10)

        if hasattr(self, "billing_quick_group"):
            self.billing_quick_group.setVisible(not compact)

        row_height = 30 if compact else 38
        self.billing_items_table.verticalHeader().setDefaultSectionSize(row_height)
        self.cart_table.verticalHeader().setDefaultSectionSize(34 if compact else 40)

        if hasattr(self, "billing_splitter"):
            self.billing_splitter.setSizes([470, 760] if compact else [520, 780])

    def _make_billing_stat_card(self, heading: str) -> tuple[QFrame, QLabel]:
        card = QFrame()
        card.setObjectName("BillingStatCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(2)

        title = QLabel(heading)
        title.setObjectName("BillingStatTitle")
        value = QLabel("0")
        value.setObjectName("BillingStatValue")

        layout.addWidget(title)
        layout.addWidget(value)
        return card, value

    def _style_billing_table(self, table: QTableWidget) -> None:
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        table.setWordWrap(False)
        table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        table.horizontalHeader().setMinimumSectionSize(72)
        table.horizontalHeader().setDefaultSectionSize(132)
        table.verticalHeader().setDefaultSectionSize(38)

    def _apply_billing_dashboard_style(self, tab: QWidget) -> None:
        tab.setStyleSheet(
            "#BillingHero {"
            "border: 1px solid #3a4457;"
            "border-radius: 8px;"
            "background-color: #222a37;"
            "}"
            "#BillingHeroTitle {"
            "font-size: 18px;"
            "font-weight: 700;"
            "color: #f2f7ff;"
            "}"
            "#BillingEmptyState {"
            "padding: 6px 10px;"
            "border: 1px dashed #57627c;"
            "border-radius: 6px;"
            "color: #f1d4a6;"
            "background-color: #2e2a20;"
            "font-weight: 600;"
            "}"
            "#BillingStatCard {"
            "border: 1px solid #3f4a60;"
            "border-radius: 8px;"
            "background-color: #1b2230;"
            "min-width: 96px;"
            "}"
            "#BillingStatTitle {"
            "font-size: 10px;"
            "color: #9eb0cb;"
            "}"
            "#BillingStatValue {"
            "font-size: 17px;"
            "font-weight: 700;"
            "color: #ffffff;"
            "}"
            "#BillingDashboard QGroupBox {"
            "border: 1px solid #3a4252;"
            "border-radius: 8px;"
            "margin-top: 8px;"
            "padding-top: 8px;"
            "}"
            "#BillingDashboard QGroupBox::title {"
            "subcontrol-origin: margin;"
            "left: 10px;"
            "padding: 0 4px;"
            "color: #dbe5f5;"
            "font-weight: 600;"
            "}"
            "#BillingDashboard QTableWidget {"
            "gridline-color: #2f4261;"
            "alternate-background-color: #1b2942;"
            "background-color: #152036;"
            "border: 1px solid #385176;"
            "border-radius: 6px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "#BillingDashboard QTableWidget::item:hover {"
            "background-color: #203a5f;"
            "}"
            "#BillingDashboard QHeaderView::section {"
            "background-color: #233552;"
            "color: #eaf1ff;"
            "padding: 7px 9px;"
            "border: 0px;"
            "border-right: 1px solid #3f5980;"
            "font-weight: 600;"
            "}"
            "#BillingDashboard QLineEdit,"
            "#BillingDashboard QDoubleSpinBox,"
            "#BillingDashboard QComboBox {"
            "background-color: #131d2e;"
            "color: #f2f6ff;"
            "border: 1px solid #425a80;"
            "border-radius: 6px;"
            "padding: 6px 8px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "#BillingDashboard QLineEdit::placeholder {"
            "color: #90a4c8;"
            "}"
            "#BillingDashboard QComboBox::drop-down {"
            "border: 0px;"
            "width: 22px;"
            "}"
            "#BillingDashboard QComboBox QAbstractItemView {"
            "background-color: #1a263c;"
            "color: #edf3ff;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #435e88;"
            "}"
            "#BillingDashboard QPushButton {"
            "background-color: #344055;"
            "color: #eef3fb;"
            "border: 1px solid #46546d;"
            "border-radius: 8px;"
            "padding: 8px 12px;"
            "}"
            "#BillingDashboard QPushButton:hover {"
            "background-color: #3e4e68;"
            "}"
            "#BillingDashboard QLineEdit:focus,"
            "#BillingDashboard QDoubleSpinBox:focus,"
            "#BillingDashboard QTableWidget:focus,"
            "#BillingDashboard QPushButton:focus,"
            "#BillingDashboard QCheckBox:focus {"
            "border: 1px solid #53b8ff;"
            "outline: none;"
            "background-color: #243047;"
            "}"
            "#BillingDashboard QPushButton#PrimaryBillingButton {"
            "background-color: #1f8b4c;"
            "border: 1px solid #25a55a;"
            "color: white;"
            "font-weight: 700;"
            "}"
            "#BillingDashboard QPushButton#PrimaryBillingButton:hover {"
            "background-color: #26a059;"
            "}"
            "#BillingDashboard QPushButton#SecondaryBillingButton {"
            "background-color: #2d3a50;"
            "border: 1px solid #4a5872;"
            "}"
            "#BillingDashboard QPushButton#SecondaryBillingButton:hover {"
            "background-color: #384964;"
            "}"
            "#BillingTotalPanel {"
            "border: 1px solid #3b465b;"
            "border-radius: 8px;"
            "background-color: #1a2738;"
            "}"
            "#BillingTotalLabel {"
            "font-size: 34px;"
            "font-weight: 800;"
            "color: #f5fbff;"
            "}"
        )

    def _apply_inventory_panel_style(self, tab: QWidget) -> None:
        tab.setStyleSheet(
            "#InventoryPanel QGroupBox {"
            "border: 1px solid #3d475b;"
            "border-radius: 8px;"
            "margin-top: 8px;"
            "padding-top: 8px;"
            "}"
            "#InventoryPanel QGroupBox::title {"
            "subcontrol-origin: margin;"
            "left: 8px;"
            "padding: 0 4px;"
            "font-weight: 600;"
            "color: #d9e3f3;"
            "}"
            "#InventoryPanel QTableWidget {"
            "gridline-color: #2f4261;"
            "background-color: #152036;"
            "alternate-background-color: #1b2942;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #385176;"
            "border-radius: 6px;"
            "}"
            "#InventoryPanel QTableWidget::item:hover {"
            "background-color: #203a5f;"
            "}"
            "#InventoryPanel QTableWidget::item:selected {"
            "background-color: #2b77e7;"
            "color: #ffffff;"
            "}"
            "#InventoryPanel QHeaderView::section {"
            "background-color: #233552;"
            "color: #eaf1ff;"
            "padding: 7px 8px;"
            "border: 0px;"
            "border-right: 1px solid #3f5980;"
            "font-weight: 600;"
            "}"
            "#InventoryPanel QLineEdit,"
            "#InventoryPanel QDoubleSpinBox,"
            "#InventoryPanel QComboBox {"
            "background-color: #131d2e;"
            "color: #f2f6ff;"
            "border: 1px solid #425a80;"
            "border-radius: 6px;"
            "padding: 6px 8px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "#InventoryPanel QLineEdit::placeholder {"
            "color: #90a4c8;"
            "}"
            "#InventoryPanel QComboBox::drop-down {"
            "border: 0px;"
            "width: 22px;"
            "}"
            "#InventoryPanel QComboBox QAbstractItemView {"
            "background-color: #1a263c;"
            "color: #edf3ff;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #435e88;"
            "}"
            "#InventoryPanel QLineEdit:focus,"
            "#InventoryPanel QDoubleSpinBox:focus,"
            "#InventoryPanel QComboBox:focus {"
            "border: 1px solid #55c2ff;"
            "background-color: #1a2940;"
            "}"
            "#InventoryPanel QPushButton, #InventoryPanel QToolButton {"
            "border-radius: 8px;"
            "padding: 7px 12px;"
            "border: 1px solid #4a5872;"
            "background-color: #2d3a50;"
            "color: #edf3fb;"
            "}"
            "#InventoryPanel QPushButton#PrimaryInventoryButton {"
            "background-color: #1f8b4c;"
            "border: 1px solid #25a55a;"
            "font-weight: 700;"
            "}"
            "#InventoryPanel QPushButton#PrimaryInventoryButton:hover {"
            "background-color: #26a059;"
            "}"
            "#InventoryInlineStatus {"
            "color: #a8f0bc;"
            "font-weight: 600;"
            "padding-left: 2px;"
            "}"
            "#InventorySummaryCard, #InventorySummaryLow, #InventorySummaryOOS {"
            "border: 1px solid #3f4b62;"
            "border-radius: 8px;"
            "background-color: #1e2838;"
            "padding: 6px 10px;"
            "font-weight: 600;"
            "color: #d9e4f7;"
            "}"
            "#InventorySummaryLow {"
            "border-color: #6f6130;"
            "background-color: #2f2a1b;"
            "color: #ffe9b2;"
            "}"
            "#InventorySummaryOOS {"
            "border-color: #6b3e3e;"
            "background-color: #2f2020;"
            "color: #ffd4d4;"
            "}"
            "#InventoryEmptyState {"
            "padding: 7px 10px;"
            "border: 1px dashed #536179;"
            "border-radius: 6px;"
            "color: #f2d7a8;"
            "background-color: #2e2a22;"
            "}"
        )

    def _apply_purchases_panel_style(self, tab: QWidget) -> None:
        tab.setStyleSheet(
            "#PurchasesPanel QGroupBox {"
            "border: 1px solid #3f485b;"
            "border-radius: 8px;"
            "margin-top: 8px;"
            "padding-top: 8px;"
            "}"
            "#PurchasesPanel QGroupBox::title {"
            "subcontrol-origin: margin;"
            "left: 8px;"
            "padding: 0 4px;"
            "font-weight: 600;"
            "color: #dbe5f5;"
            "}"
            "#PurchasesPanel QTableWidget {"
            "gridline-color: #2f4261;"
            "background-color: #152036;"
            "alternate-background-color: #1b2942;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #385176;"
            "border-radius: 6px;"
            "}"
            "#PurchasesPanel QTableWidget::item:hover {"
            "background-color: #203a5f;"
            "}"
            "#PurchasesPanel QHeaderView::section {"
            "background-color: #233552;"
            "color: #eaf1ff;"
            "padding: 7px 8px;"
            "border: 0px;"
            "border-right: 1px solid #3f5980;"
            "font-weight: 600;"
            "}"
            "#PurchasesPanel QLineEdit,"
            "#PurchasesPanel QDoubleSpinBox,"
            "#PurchasesPanel QComboBox,"
            "#PurchasesPanel QDateEdit {"
            "background-color: #131d2e;"
            "color: #f2f6ff;"
            "border: 1px solid #425a80;"
            "border-radius: 6px;"
            "padding: 6px 8px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "#PurchasesPanel QLineEdit::placeholder {"
            "color: #90a4c8;"
            "}"
            "#PurchasesPanel QComboBox::drop-down,"
            "#PurchasesPanel QDateEdit::drop-down {"
            "border: 0px;"
            "width: 22px;"
            "}"
            "#PurchasesPanel QComboBox QAbstractItemView,"
            "#PurchasesPanel QDateEdit QAbstractItemView {"
            "background-color: #1a263c;"
            "color: #edf3ff;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #435e88;"
            "}"
            "#PurchasesPanel QLineEdit:focus,"
            "#PurchasesPanel QDoubleSpinBox:focus,"
            "#PurchasesPanel QComboBox:focus,"
            "#PurchasesPanel QDateEdit:focus {"
            "border: 1px solid #55c2ff;"
            "background-color: #1a2940;"
            "}"
            "#PurchasesPanel QPushButton {"
            "border-radius: 8px;"
            "padding: 7px 12px;"
            "border: 1px solid #4a5872;"
            "background-color: #2d3a50;"
            "color: #edf3fb;"
            "}"
            "#PurchasesPanel QPushButton#PrimaryPurchaseButton {"
            "background-color: #219552;"
            "border: 1px solid #25a55a;"
            "font-weight: 700;"
            "padding: 10px 16px;"
            "}"
            "#PurchasesPanel QPushButton#PrimaryPurchaseButton:hover {"
            "background-color: #2bb061;"
            "border-color: #6fe3a5;"
            "}"
            "#PurchasesPanel QPushButton#SecondaryPurchaseButton {"
            "background-color: #2d3a50;"
            "}"
            "#PurchasesPanel QPushButton#PurchaseFilterPreset:checked {"
            "background-color: #3b4f72;"
            "border-color: #6b87b8;"
            "font-weight: 700;"
            "}"
            "#PurchaseTotalPanel {"
            "border: 1px solid #3a465b;"
            "border-radius: 8px;"
            "background-color: qlineargradient(x1:0,y1:0,x2:1,y2:1,stop:0 #1e2b3e, stop:1 #1a2433);"
            "}"
            "#PurchaseTotalLabel {"
            "font-size: 18px;"
            "font-weight: 800;"
            "color: #f4fbff;"
            "}"
            "#PurchaseEmptyState {"
            "padding: 7px 10px;"
            "border: 1px dashed #55627a;"
            "border-radius: 6px;"
            "background-color: #2c2a21;"
            "color: #f2d9ab;"
            "}"
            "#PurchaseInlineFeedback {"
            "color: #a9b8d1;"
            "font-weight: 600;"
            "}"
            "#PurchaseStockPreview {"
            "padding: 4px 8px;"
            "border: 1px solid #446448;"
            "border-radius: 6px;"
            "background-color: #223626;"
            "color: #b7f0c2;"
            "font-weight: 700;"
            "}"
        )

    def _show_selected_purchase_details(self, _item: QTableWidgetItem | None = None) -> None:
        selected = self.purchase_history_table.currentRow()
        if selected < 0:
            return

        supplier = self.purchase_history_table.item(selected, 1).text() if self.purchase_history_table.item(selected, 1) else "-"
        date_text = self.purchase_history_table.item(selected, 2).text() if self.purchase_history_table.item(selected, 2) else "-"
        lines = self.purchase_history_table.item(selected, 3).text() if self.purchase_history_table.item(selected, 3) else "0"
        total = self.purchase_history_table.item(selected, 4).text() if self.purchase_history_table.item(selected, 4) else "0.00"
        notes = self.purchase_history_table.item(selected, 5).text() if self.purchase_history_table.item(selected, 5) else ""

        QMessageBox.information(
            self,
            "Purchase Details",
            f"Supplier: {supplier}\nDate: {date_text}\nLines: {lines}\nTotal: INR {total}\nNotes: {notes or '-'}",
        )

    def _set_purchase_filter_button_state(self, preset: str) -> None:
        buttons = [
            getattr(self, "purchase_filter_today_btn", None),
            getattr(self, "purchase_filter_7d_btn", None),
            getattr(self, "purchase_filter_month_btn", None),
            getattr(self, "purchase_filter_custom_btn", None),
        ]
        for btn in buttons:
            if btn is not None:
                btn.setChecked(False)

        mapping = {
            "today": getattr(self, "purchase_filter_today_btn", None),
            "last7": getattr(self, "purchase_filter_7d_btn", None),
            "month": getattr(self, "purchase_filter_month_btn", None),
            "custom": getattr(self, "purchase_filter_custom_btn", None),
        }
        active_btn = mapping.get(preset)
        if active_btn is not None:
            active_btn.setChecked(True)

    def _set_purchase_filter_preset(self, preset: str) -> None:
        if preset in ("today", "last7", "month"):
            self._apply_quick_range(
                self.purchase_from_date,
                self.purchase_to_date,
                preset,
                self.refresh_purchases_tab,
            )
        else:
            self.refresh_purchases_tab()
        self._set_purchase_filter_button_state(preset)

    def duplicate_selected_purchase(self) -> None:
        selected = self.purchase_history_table.currentRow()
        if selected < 0:
            QMessageBox.information(self, "Duplicate Purchase", "Please select a purchase from history.")
            return

        id_cell = self.purchase_history_table.item(selected, 0)
        if id_cell is None:
            return

        purchase_id = int(id_cell.text())
        pin = self._require_admin_access("Duplicate Purchase")
        if pin is None:
            return

        try:
            purchase = self.bookkeeping_service.get_purchase_for_edit(purchase_id=purchase_id, admin_pin=pin)
        except ValueError as exc:
            QMessageBox.warning(self, "Duplicate Purchase", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Duplicate Purchase", str(exc))
            return

        self.purchase_supplier_input.setText(purchase.get("supplier_name") or "")
        self.purchase_notes_input.setText(purchase.get("notes") or "")
        self.purchase_cart = [
            {
                "item_id": int(line["item_id"]),
                "name": line["name"],
                "quantity": float(line["quantity"]),
                "cost_price": float(line["cost_price"]),
            }
            for line in purchase.get("items", [])
        ]
        self.refresh_purchase_lines_table()
        self._set_purchase_mode(False)
        if hasattr(self, "purchase_feedback_label"):
            self.purchase_feedback_label.setText(f"Purchase #{purchase_id} duplicated into builder.")
            self.purchase_feedback_label.setVisible(True)
            QTimer.singleShot(2200, lambda: self.purchase_feedback_label.setVisible(False))

    def _sync_inventory_add_button_state(self) -> None:
        if not hasattr(self, "inventory_add_btn"):
            return
        is_valid = (
            bool(self.item_name_input.text().strip())
            and float(self.sell_price_spin.value()) > 0
            and float(self.stock_spin.value()) >= 0
            and float(self.reorder_spin.value()) >= 0
        )
        self.inventory_add_btn.setEnabled(is_valid)

    def _on_inventory_edit_started(self, item: QTableWidgetItem) -> None:
        item.setBackground(QColor(50, 70, 108))
        if hasattr(self, "inventory_inline_status_label"):
            self.inventory_inline_status_label.setText("Editing inline... press Enter to save.")
            self.inventory_inline_status_label.setVisible(True)

    def _inventory_id_for_row(self, row: int) -> int | None:
        if row < 0:
            return None
        name_cell = self.inventory_items_table.item(row, 0)
        if name_cell is None:
            return None
        item_id = name_cell.data(Qt.UserRole)
        if item_id is None:
            return None
        return int(item_id)

    def _update_low_stock_panel(self, items: list[dict]) -> None:
        if not hasattr(self, "low_stock_list"):
            return

        self.low_stock_list.clear()
        low_stock_items = [
            i for i in items if float(i.get("stock_quantity", 0)) <= float(i.get("reorder_level", 0))
        ]
        low_stock_items.sort(key=lambda x: float(x.get("stock_quantity", 0)))

        if not low_stock_items:
            self.low_stock_list.addItem(QListWidgetItem("No low stock alerts"))
            return

        grouped: dict[str, dict[str, float]] = {}
        for item in low_stock_items:
            name = item.get("name", "-")
            stock = float(item.get("stock_quantity", 0))
            current = grouped.get(name)
            if current is None:
                grouped[name] = {"count": 1, "min_stock": stock}
            else:
                current["count"] += 1
                current["min_stock"] = min(current["min_stock"], stock)

        sorted_rows = sorted(grouped.items(), key=lambda kv: (-kv[1]["count"], kv[0].lower()))
        for name, info in sorted_rows[:12]:
            if info["count"] > 1:
                status = f"{name} -> {int(info['count'])} variants low"
            else:
                status = f"{name} ({info['min_stock']:.2f} left)"
            entry = QListWidgetItem(status)
            entry.setData(Qt.UserRole, name)
            self.low_stock_list.addItem(entry)

    def _on_low_stock_item_clicked(self, item: QListWidgetItem) -> None:
        name = item.data(Qt.UserRole)
        if not name:
            return
        self.inventory_search_input.setText(str(name))
        self.inventory_low_stock_only_checkbox.setChecked(True)

    def _update_inventory_summary(self, items: list[dict]) -> None:
        if not hasattr(self, "inventory_summary_total"):
            return

        total_count = len(items)
        low_stock_count = 0
        out_of_stock_count = 0
        for item in items:
            stock = float(item.get("stock_quantity", 0))
            reorder = float(item.get("reorder_level", 0))
            if stock <= 0:
                out_of_stock_count += 1
            if stock <= reorder:
                low_stock_count += 1

        self.inventory_summary_total.setText(f"Total Items: {total_count}")
        self.inventory_summary_low.setText(f"Low Stock: {low_stock_count}")
        self.inventory_summary_oos.setText(f"Out of Stock: {out_of_stock_count}")

    def apply_inventory_filter(self) -> None:
        if not hasattr(self, "inventory_items_cache"):
            return

        query = self.inventory_search_input.text().strip().lower()
        category_id = self.inventory_filter_category_combo.currentData()
        low_stock_only = self.inventory_low_stock_only_checkbox.isChecked()

        filtered: list[dict] = []
        for item in self.inventory_items_cache:
            name = item.get("name", "")
            item_category_id = item.get("category_id")
            stock = float(item.get("stock_quantity", 0))
            reorder = float(item.get("reorder_level", 0))

            if query and query not in name.lower():
                continue
            if category_id is not None and int(item_category_id or -1) != int(category_id):
                continue
            if low_stock_only and stock > reorder:
                continue
            filtered.append(item)

        table = self.inventory_items_table
        self._updating_inventory_table = True
        table.setSortingEnabled(False)
        table.setRowCount(len(filtered))

        for row_index, item in enumerate(filtered):
            name_item = QTableWidgetItem(item["name"])
            name_item.setData(Qt.UserRole, int(item["id"]))
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_index, 0, name_item)

            category_item = QTableWidgetItem(item.get("category_name") or "-")
            category_item.setFlags(category_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_index, 1, category_item)

            kind_text = "Ingredient" if (item.get("item_kind") or "sellable") == "ingredient" else "Sellable"
            kind_item = QTableWidgetItem(kind_text)
            kind_item.setFlags(kind_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_index, 2, kind_item)

            costing_raw = item.get("costing_mode") or "manual"
            costing_text = "Recipe" if costing_raw == "recipe" else "Manual"
            costing_item = QTableWidgetItem(costing_text)
            costing_item.setFlags(costing_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_index, 3, costing_item)

            sell_item = QTableWidgetItem(f"{float(item['selling_price']):.2f}")
            sell_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            table.setItem(row_index, 4, sell_item)

            stock_value = float(item["stock_quantity"])
            reorder_value = float(item["reorder_level"])
            stock_item = QTableWidgetItem(f"{stock_value:.2f}")
            stock_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            stock_item.setFlags(stock_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_index, 5, stock_item)

            reorder_item = QTableWidgetItem(f"{reorder_value:.2f}")
            reorder_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            table.setItem(row_index, 6, reorder_item)

            actions_cell = QWidget()
            actions_layout = QHBoxLayout(actions_cell)
            actions_layout.setContentsMargins(2, 1, 2, 1)
            actions_layout.setSpacing(4)

            restock_btn = QToolButton()
            restock_btn.setText("+Stock")
            restock_btn.clicked.connect(lambda _, item_id=int(item["id"]): self._quick_restock_inventory_item(item_id))

            edit_btn = QToolButton()
            edit_btn.setText("Edit")
            edit_btn.clicked.connect(lambda _, item_id=int(item["id"]): self._edit_inventory_item_by_id(item_id))

            delete_btn = QToolButton()
            delete_btn.setText("Delete")
            delete_btn.clicked.connect(lambda _, item_id=int(item["id"]): self._delete_inventory_item_by_id(item_id))

            actions_layout.addWidget(restock_btn)
            actions_layout.addWidget(edit_btn)
            actions_layout.addWidget(delete_btn)
            actions_layout.addStretch()
            table.setCellWidget(row_index, 7, actions_cell)

            if stock_value <= 0:
                stock_item.setBackground(QColor("#3a1f1f"))
                stock_item.setForeground(QColor(255, 232, 232))
            elif stock_value < reorder_value:
                stock_item.setBackground(QColor("#3a3420"))
                stock_item.setForeground(QColor(255, 245, 204))

        table.setSortingEnabled(True)

        self._updating_inventory_table = False
        if hasattr(self, "inventory_empty_state_label"):
            self.inventory_empty_state_label.setVisible(len(filtered) == 0)
        self._update_inventory_summary(self.inventory_items_cache)
        self._update_low_stock_panel(self.inventory_items_cache)

    def _select_inventory_row_by_item_id(self, item_id: int) -> bool:
        for row in range(self.inventory_items_table.rowCount()):
            row_item_id = self._inventory_id_for_row(row)
            if row_item_id == int(item_id):
                self.inventory_items_table.selectRow(row)
                self.inventory_items_table.scrollToItem(self.inventory_items_table.item(row, 0))
                return True
        return False

    def _edit_inventory_item_by_id(self, item_id: int) -> None:
        if self._select_inventory_row_by_item_id(item_id):
            self.update_selected_item_price()

    def _delete_inventory_item_by_id(self, item_id: int) -> None:
        if self._select_inventory_row_by_item_id(item_id):
            self.delete_selected_item()

    def _quick_restock_inventory_item(self, item_id: int) -> None:
        if not self._select_inventory_row_by_item_id(item_id):
            return

        item = self._selected_inventory_item()
        if item is None:
            return

        pin = self._require_admin_access("Quick Restock")
        if pin is None:
            return

        add_qty, ok_qty = QInputDialog.getDouble(
            self,
            "Quick Restock",
            f"Add stock quantity for {item['name']}:",
            value=1.0,
            minValue=0.01,
            decimals=2,
        )
        if not ok_qty:
            return

        try:
            self.inventory_service.manual_stock_adjustment(
                item_id=item["item_id"],
                quantity_delta=float(add_qty),
                admin_pin=pin,
                notes="Quick restock from inventory table",
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Quick Restock", str(exc))
            return

        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_reports()
        self.inventory_inline_status_label.setText(f"Stock added: {item['name']} +{add_qty:.2f}")
        self.inventory_inline_status_label.setVisible(True)
        QTimer.singleShot(2200, lambda: self.inventory_inline_status_label.setVisible(False))

    def _build_inventory_tab(self) -> QWidget:
        tab = QWidget()
        tab.setObjectName("InventoryPanel")
        root_layout = QVBoxLayout(tab)
        root_layout.setContentsMargins(10, 10, 10, 10)
        root_layout.setSpacing(8)

        form_group = QGroupBox("Add Item")
        form_layout = QGridLayout(form_group)
        form_layout.setHorizontalSpacing(10)
        form_layout.setVerticalSpacing(8)

        basic_group = QGroupBox("Basic Info")
        basic_layout = QGridLayout(basic_group)

        pricing_group = QGroupBox("Pricing")
        pricing_layout = QGridLayout(pricing_group)

        stock_group = QGroupBox("Stock")
        stock_layout = QGridLayout(stock_group)

        self.item_name_input = QLineEdit()
        self.item_name_input.setPlaceholderText("Item name")
        self.item_name_input.returnPressed.connect(self.add_inventory_item)
        self.category_combo = QComboBox()
        self.item_kind_combo = QComboBox()
        self.item_kind_combo.addItem("Sellable", "sellable")
        self.item_kind_combo.addItem("Ingredient", "ingredient")
        self.costing_mode_combo = QComboBox()
        self.costing_mode_combo.addItem("Manual Cost", "manual")
        self.costing_mode_combo.addItem("Recipe Cost", "recipe")
        self.unit_name_input = QLineEdit("pcs")
        self.unit_name_input.setPlaceholderText("Unit (pcs, g, ml, etc.)")
        self.stock_tracked_checkbox = QCheckBox("Track stock")
        self.stock_tracked_checkbox.setChecked(True)

        self.sell_price_spin = QDoubleSpinBox()
        self.sell_price_spin.setMaximum(100000)
        self.sell_price_spin.setPrefix("INR ")

        self.cost_price_spin = QDoubleSpinBox()
        self.cost_price_spin.setMaximum(100000)
        self.cost_price_spin.setPrefix("INR ")

        self.stock_spin = QDoubleSpinBox()
        self.stock_spin.setMaximum(100000)

        self.reorder_spin = QDoubleSpinBox()
        self.reorder_spin.setMaximum(100000)

        self.inventory_add_btn = QPushButton("+ Add Item")
        self.inventory_add_btn.setObjectName("PrimaryInventoryButton")
        self.inventory_add_btn.setMinimumHeight(40)
        self.inventory_add_btn.setMinimumWidth(126)
        self.inventory_add_btn.clicked.connect(self.add_inventory_item)

        basic_layout.addWidget(QLabel("<b>Name</b>"), 0, 0)
        basic_layout.addWidget(self.item_name_input, 0, 1)
        basic_layout.addWidget(QLabel("<b>Category</b>"), 1, 0)
        basic_layout.addWidget(self.category_combo, 1, 1)
        basic_layout.addWidget(QLabel("<b>Kind</b>"), 2, 0)
        basic_layout.addWidget(self.item_kind_combo, 2, 1)
        basic_layout.addWidget(QLabel("<b>Unit</b>"), 3, 0)
        basic_layout.addWidget(self.unit_name_input, 3, 1)

        pricing_layout.addWidget(QLabel("<b>Selling Price</b>"), 0, 0)
        pricing_layout.addWidget(self.sell_price_spin, 0, 1)
        pricing_layout.addWidget(QLabel("<b>Cost Price</b>"), 1, 0)
        pricing_layout.addWidget(self.cost_price_spin, 1, 1)
        pricing_layout.addWidget(QLabel("<b>Costing</b>"), 2, 0)
        pricing_layout.addWidget(self.costing_mode_combo, 2, 1)
        pricing_layout.addWidget(self.stock_tracked_checkbox, 3, 1)

        stock_layout.addWidget(QLabel("<b>Opening Stock</b>"), 0, 0)
        stock_layout.addWidget(self.stock_spin, 0, 1)
        stock_layout.addWidget(QLabel("<b>Reorder Level</b>"), 1, 0)
        stock_layout.addWidget(self.reorder_spin, 1, 1)

        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)

        add_row = QHBoxLayout()
        add_row.setContentsMargins(0, 4, 0, 0)
        add_row.addStretch()
        add_row.addWidget(self.inventory_add_btn)

        form_layout.addWidget(basic_group, 0, 0)
        form_layout.addWidget(pricing_group, 0, 1)
        form_layout.addWidget(stock_group, 0, 2)
        form_layout.addWidget(divider, 1, 0, 1, 3)
        form_layout.addLayout(add_row, 2, 0, 1, 3)

        inventory_hint = QLabel(
            "Tip: Double-click Sell/Reorder to edit inline. Red=out of stock, Yellow=below reorder."
        )

        filter_row = QHBoxLayout()
        filter_row.setContentsMargins(0, 0, 0, 0)
        filter_row.setSpacing(8)
        self.inventory_search_input = QLineEdit()
        self.inventory_search_input.setPlaceholderText("Search item (name, category)...")
        self.inventory_search_input.setClearButtonEnabled(True)
        self.inventory_search_input.setMinimumHeight(34)
        self.inventory_search_input.textChanged.connect(self.apply_inventory_filter)
        self.inventory_filter_category_combo = QComboBox()
        self.inventory_filter_category_combo.addItem("All Categories", None)
        self.inventory_filter_category_combo.setMinimumWidth(220)
        self.inventory_filter_category_combo.setMinimumHeight(34)
        self.inventory_filter_category_combo.currentIndexChanged.connect(self.apply_inventory_filter)
        self.inventory_low_stock_only_checkbox = QCheckBox("Low Stock Only")
        self.inventory_low_stock_only_checkbox.toggled.connect(self.apply_inventory_filter)
        filter_row.addWidget(self.inventory_search_input, 2)
        filter_row.addWidget(QLabel("<b>Category</b>"))
        filter_row.addWidget(self.inventory_filter_category_combo, 1)
        filter_row.addWidget(self.inventory_low_stock_only_checkbox)

        summary_row = QHBoxLayout()
        self.inventory_summary_total = QLabel("Total Items: 0")
        self.inventory_summary_total.setObjectName("InventorySummaryCard")
        self.inventory_summary_low = QLabel("Low Stock: 0")
        self.inventory_summary_low.setObjectName("InventorySummaryLow")
        self.inventory_summary_oos = QLabel("Out of Stock: 0")
        self.inventory_summary_oos.setObjectName("InventorySummaryOOS")
        summary_row.addWidget(self.inventory_summary_total)
        summary_row.addWidget(self.inventory_summary_low)
        summary_row.addWidget(self.inventory_summary_oos)
        summary_row.addStretch()

        table_area_layout = QHBoxLayout()
        table_area_layout.setSpacing(10)

        self.inventory_items_table = QTableWidget(0, 8)
        self.inventory_items_table.setHorizontalHeaderLabels(
            ["Name", "Category", "Kind", "Costing", "Sell", "Stock", "Reorder", "Actions"]
        )
        self.inventory_items_table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.SelectedClicked | QTableWidget.EditKeyPressed
        )
        self.inventory_items_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.inventory_items_table.itemChanged.connect(self._on_inventory_item_changed)
        self.inventory_items_table.itemDoubleClicked.connect(self._on_inventory_edit_started)
        self.inventory_items_table.setSortingEnabled(True)
        self.inventory_items_table.horizontalHeader().setStretchLastSection(False)
        self.inventory_items_table.horizontalHeader().setMinimumSectionSize(90)
        self.inventory_items_table.verticalHeader().setDefaultSectionSize(38)
        self.inventory_items_table.setAlternatingRowColors(True)
        header = self.inventory_items_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.inventory_items_table.horizontalHeaderItem(4).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.inventory_items_table.horizontalHeaderItem(5).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.inventory_items_table.horizontalHeaderItem(6).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

        low_stock_group = QGroupBox("Low Stock Items")
        low_stock_layout = QVBoxLayout(low_stock_group)
        self.low_stock_list = QListWidget()
        self.low_stock_list.itemClicked.connect(self._on_low_stock_item_clicked)
        low_stock_layout.addWidget(self.low_stock_list)

        table_area_layout.addWidget(self.inventory_items_table, 4)
        table_area_layout.addWidget(low_stock_group, 1)

        self.inventory_inline_status_label = QLabel("")
        self.inventory_inline_status_label.setObjectName("InventoryInlineStatus")
        self.inventory_inline_status_label.setVisible(False)

        self.inventory_empty_state_label = QLabel("No items found. Click '+ Add Item' to get started.")
        self.inventory_empty_state_label.setObjectName("InventoryEmptyState")
        self.inventory_empty_state_label.setVisible(False)

        action_row = QHBoxLayout()
        refresh_btn = QPushButton("Refresh Inventory")
        refresh_btn.setObjectName("PrimaryInventoryButton")
        refresh_btn.clicked.connect(self.refresh_inventory)

        actions_menu_btn = QToolButton()
        actions_menu_btn.setText("Inventory Actions")
        actions_menu_btn.setPopupMode(QToolButton.InstantPopup)
        actions_menu = QMenu(actions_menu_btn)
        actions_menu.addAction("Export CSV", self.export_inventory_csv)
        actions_menu.addAction("Update Price (Admin PIN)", self.update_selected_item_price)
        actions_menu.addAction("Manual Stock Adjust (Admin PIN)", self.adjust_selected_item_stock)
        actions_menu.addAction("Manage Recipe (Admin PIN)", self.manage_selected_item_recipe)
        actions_menu.addAction("Delete Item (Admin PIN)", self.delete_selected_item)
        actions_menu.addAction("Load Starter Cigarette SKUs", self.load_starter_cigarettes)
        actions_menu_btn.setMenu(actions_menu)

        action_row.addWidget(refresh_btn)
        action_row.addWidget(actions_menu_btn)
        action_row.addStretch()

        root_layout.addWidget(form_group)
        root_layout.addLayout(filter_row)
        root_layout.addLayout(summary_row)
        root_layout.addWidget(inventory_hint)
        root_layout.addLayout(table_area_layout)
        root_layout.addWidget(self.inventory_empty_state_label)
        root_layout.addWidget(self.inventory_inline_status_label)
        root_layout.addLayout(action_row)
        root_layout.setStretch(5, 1)

        self.item_name_input.textChanged.connect(self._sync_inventory_add_button_state)
        self.sell_price_spin.valueChanged.connect(self._sync_inventory_add_button_state)
        self.stock_spin.valueChanged.connect(self._sync_inventory_add_button_state)
        self.reorder_spin.valueChanged.connect(self._sync_inventory_add_button_state)
        self.category_combo.currentIndexChanged.connect(self._sync_inventory_add_button_state)
        self.item_kind_combo.currentIndexChanged.connect(self._on_item_kind_changed)
        self._on_item_kind_changed()
        self._sync_inventory_add_button_state()
        self._apply_inventory_panel_style(tab)

        return tab

    def _build_purchases_tab(self) -> QWidget:
        tab = QWidget()
        tab.setObjectName("PurchasesPanel")
        root_layout = QVBoxLayout(tab)
        root_layout.setContentsMargins(10, 10, 10, 10)
        root_layout.setSpacing(4)

        entry_box = QGroupBox("Purchase Builder")
        entry_layout = QGridLayout(entry_box)
        entry_layout.setHorizontalSpacing(10)
        entry_layout.setVerticalSpacing(8)

        self.purchase_supplier_input = QLineEdit()
        self.purchase_supplier_input.setPlaceholderText("Supplier name")
        self.purchase_notes_input = QLineEdit()
        self.purchase_notes_input.setPlaceholderText("Optional notes")
        self.purchase_item_combo = QComboBox()
        self.purchase_item_combo.currentIndexChanged.connect(self._on_purchase_item_changed)
        self.purchase_qty_spin = QDoubleSpinBox()
        self.purchase_qty_spin.setDecimals(2)
        self.purchase_qty_spin.setMinimum(0.01)
        self.purchase_qty_spin.setMaximum(100000)
        self.purchase_qty_spin.setValue(1)
        self.purchase_qty_spin.valueChanged.connect(self._update_purchase_stock_preview)

        self.purchase_cost_spin = QDoubleSpinBox()
        self.purchase_cost_spin.setDecimals(2)
        self.purchase_cost_spin.setMinimum(0)
        self.purchase_cost_spin.setMaximum(100000)
        self.purchase_cost_spin.setPrefix("INR ")
        self.purchase_item_combo.setMaximumWidth(460)
        self.purchase_qty_spin.setMaximumWidth(130)
        self.purchase_cost_spin.setMaximumWidth(180)

        self.add_line_btn = QPushButton("+ Add Item to Purchase")
        self.add_line_btn.setObjectName("PrimaryPurchaseButton")
        self.add_line_btn.setMinimumHeight(46)
        self.add_line_btn.setMinimumWidth(280)
        self.add_line_btn.clicked.connect(self.add_purchase_line)

        save_purchase_btn = QPushButton("Save Purchase")
        save_purchase_btn.setObjectName("PrimaryPurchaseButton")
        save_purchase_btn.setMinimumHeight(34)
        save_purchase_btn.setMaximumHeight(34)
        self.save_purchase_btn = save_purchase_btn
        save_purchase_btn.clicked.connect(self.save_purchase)

        cancel_edit_btn = QPushButton("Cancel Edit")
        cancel_edit_btn.setObjectName("SecondaryPurchaseButton")
        self.cancel_purchase_edit_btn = cancel_edit_btn
        self.cancel_purchase_edit_btn.setVisible(False)
        self.cancel_purchase_edit_btn.setEnabled(False)
        cancel_edit_btn.clicked.connect(self.cancel_purchase_edit)

        self.clear_purchase_lines_btn = QPushButton("Clear All")
        self.clear_purchase_lines_btn.setObjectName("SecondaryPurchaseButton")
        self.clear_purchase_lines_btn.clicked.connect(self.clear_purchase_lines)

        self.remove_purchase_line_btn = QPushButton("Remove Selected")
        self.remove_purchase_line_btn.setObjectName("SecondaryPurchaseButton")
        self.remove_purchase_line_btn.clicked.connect(self.remove_selected_purchase_line)

        modify_saved_btn = QPushButton("Modify Selected Purchase (Admin PIN)")
        modify_saved_btn.setObjectName("SecondaryPurchaseButton")
        modify_saved_btn.clicked.connect(self.load_selected_purchase_for_edit)

        duplicate_saved_btn = QPushButton("Duplicate Selected Purchase")
        duplicate_saved_btn.setObjectName("SecondaryPurchaseButton")
        duplicate_saved_btn.clicked.connect(self.duplicate_selected_purchase)

        entry_layout.addWidget(QLabel("<b>Supplier</b>"), 0, 0)
        entry_layout.addWidget(self.purchase_supplier_input, 0, 1, 1, 2)
        entry_layout.addWidget(QLabel("<b>Notes</b>"), 0, 3)
        entry_layout.addWidget(self.purchase_notes_input, 0, 4, 1, 2)

        entry_layout.addWidget(QLabel("<b>Item</b>"), 1, 0)
        entry_layout.addWidget(self.purchase_item_combo, 1, 1, 1, 2)
        entry_layout.addWidget(QLabel("<b>Qty</b>"), 1, 3)
        entry_layout.addWidget(self.purchase_qty_spin, 1, 4)

        entry_layout.addWidget(QLabel("<b>Cost Price</b>"), 2, 0)
        entry_layout.addWidget(self.purchase_cost_spin, 2, 1)

        self.purchase_stock_preview_label = QLabel("")
        self.purchase_stock_preview_label.setObjectName("PurchaseStockPreview")
        entry_layout.addWidget(self.purchase_stock_preview_label, 2, 2, 1, 3)

        add_line_row = QHBoxLayout()
        add_line_row.addStretch()
        add_line_row.addWidget(self.add_line_btn)
        entry_layout.addLayout(add_line_row, 3, 0, 1, 6)

        purchase_builder_hint = QLabel(
            "Step 1: Select item and add line. Step 2: Build list. Step 3: Save purchase."
        )
        purchase_builder_hint.setObjectName("PurchaseInlineFeedback")
        purchase_builder_hint.setMaximumHeight(20)

        self.purchase_lines_table = QTableWidget(0, 4)
        self.purchase_lines_table.setHorizontalHeaderLabels(
            ["Name", "Qty", "Cost", "Line Total"]
        )
        self.purchase_lines_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.purchase_lines_table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.SelectedClicked | QTableWidget.EditKeyPressed
        )
        self.purchase_lines_table.setAlternatingRowColors(True)
        self.purchase_lines_table.itemChanged.connect(self._on_purchase_line_item_changed)
        self.purchase_lines_table.setSortingEnabled(True)
        self.purchase_lines_table.horizontalHeader().setStretchLastSection(True)
        self.purchase_lines_table.verticalHeader().setDefaultSectionSize(36)
        self.purchase_lines_table.setMinimumHeight(230)
        self.purchase_lines_table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.purchase_lines_table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.purchase_lines_table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

        self.purchase_empty_state_label = QLabel(
            "Start building your purchase. Select item above and click '+ Add Item to Purchase'."
        )
        self.purchase_empty_state_label.setObjectName("PurchaseEmptyState")
        self.purchase_empty_state_label.setVisible(True)

        table_actions_row = QHBoxLayout()
        table_actions_row.setContentsMargins(0, 0, 0, 0)
        table_actions_row.setSpacing(8)
        table_actions_row.addWidget(self.remove_purchase_line_btn)
        table_actions_row.addWidget(self.clear_purchase_lines_btn)
        table_actions_row.addStretch()

        total_panel = QFrame()
        total_panel.setObjectName("PurchaseTotalPanel")
        total_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        total_panel.setMinimumHeight(74)
        total_panel.setMaximumHeight(86)
        total_layout = QVBoxLayout(total_panel)
        total_layout.setContentsMargins(10, 4, 10, 4)
        total_layout.setSpacing(2)

        self.purchase_total_label = QLabel("TOTAL: INR 0.00")
        self.purchase_total_label.setObjectName("PurchaseTotalLabel")
        self.purchase_total_label.setAlignment(Qt.AlignCenter)

        actions_row = QHBoxLayout()
        actions_row.addStretch()
        actions_row.addWidget(self.cancel_purchase_edit_btn)
        actions_row.addWidget(save_purchase_btn)

        total_layout.addWidget(self.purchase_total_label)
        total_layout.addLayout(actions_row)

        self.purchase_mode_label = QLabel("Mode: New Purchase")
        self.purchase_mode_label.setObjectName("PurchaseInlineFeedback")
        self.purchase_mode_label.setMaximumHeight(20)
        self.purchase_feedback_label = QLabel("")
        self.purchase_feedback_label.setObjectName("PurchaseInlineFeedback")
        self.purchase_feedback_label.setVisible(False)
        purchase_hint = QLabel("Tip: Double-click Qty/Cost in a line to edit inline before saving purchase.")
        purchase_hint.setObjectName("PurchaseInlineFeedback")
        purchase_hint.setMaximumHeight(20)

        history_box = QGroupBox("Recent Purchases")
        history_layout = QVBoxLayout(history_box)

        history_filter_row = QHBoxLayout()
        self.purchase_history_table = QTableWidget(0, 6)
        self.purchase_history_table.setHorizontalHeaderLabels(
            ["ID", "Supplier", "Date", "Lines", "Total", "Notes"]
        )
        self.purchase_history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.purchase_history_table.setSortingEnabled(True)
        self.purchase_history_table.setAlternatingRowColors(True)
        self.purchase_history_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.purchase_history_table.setMinimumHeight(230)
        self.purchase_history_table.itemDoubleClicked.connect(self._show_selected_purchase_details)
        self.purchase_history_table.horizontalHeader().setStretchLastSection(True)
        self.purchase_history_table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.purchase_history_table.horizontalHeaderItem(4).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

        history_actions = QHBoxLayout()
        history_actions.addWidget(modify_saved_btn)
        history_actions.addWidget(duplicate_saved_btn)
        history_actions.addStretch()

        self.purchase_from_date = QDateEdit()
        self.purchase_from_date.setCalendarPopup(True)
        self.purchase_from_date.setDisplayFormat("yyyy-MM-dd")
        self.purchase_from_date.setDate(QDate.currentDate())

        self.purchase_to_date = QDateEdit()
        self.purchase_to_date.setCalendarPopup(True)
        self.purchase_to_date.setDisplayFormat("yyyy-MM-dd")
        self.purchase_to_date.setDate(QDate.currentDate())

        self.purchase_filter_today_btn = QPushButton("Today")
        self.purchase_filter_today_btn.setObjectName("PurchaseFilterPreset")
        self.purchase_filter_today_btn.setCheckable(True)
        self.purchase_filter_today_btn.clicked.connect(lambda: self._set_purchase_filter_preset("today"))

        self.purchase_filter_7d_btn = QPushButton("Last 7 Days")
        self.purchase_filter_7d_btn.setObjectName("PurchaseFilterPreset")
        self.purchase_filter_7d_btn.setCheckable(True)
        self.purchase_filter_7d_btn.clicked.connect(lambda: self._set_purchase_filter_preset("last7"))

        self.purchase_filter_month_btn = QPushButton("This Month")
        self.purchase_filter_month_btn.setObjectName("PurchaseFilterPreset")
        self.purchase_filter_month_btn.setCheckable(True)
        self.purchase_filter_month_btn.clicked.connect(lambda: self._set_purchase_filter_preset("month"))

        self.purchase_filter_custom_btn = QPushButton("Custom")
        self.purchase_filter_custom_btn.setObjectName("PurchaseFilterPreset")
        self.purchase_filter_custom_btn.setCheckable(True)
        self.purchase_filter_custom_btn.clicked.connect(lambda: self._set_purchase_filter_preset("custom"))

        p_apply_btn = QPushButton("Apply")
        p_apply_btn.setObjectName("PurchaseFilterPreset")
        p_apply_btn.clicked.connect(lambda: self._set_purchase_filter_preset("custom"))

        history_filter_row.addWidget(QLabel("<b>From</b>"))
        history_filter_row.addWidget(self.purchase_from_date)
        history_filter_row.addWidget(QLabel("<b>To</b>"))
        history_filter_row.addWidget(self.purchase_to_date)
        history_filter_row.addWidget(self.purchase_filter_today_btn)
        history_filter_row.addWidget(self.purchase_filter_7d_btn)
        history_filter_row.addWidget(self.purchase_filter_month_btn)
        history_filter_row.addWidget(self.purchase_filter_custom_btn)
        history_filter_row.addWidget(p_apply_btn)
        history_filter_row.addStretch()

        history_layout.addLayout(history_filter_row)
        history_layout.addWidget(self.purchase_history_table)
        history_layout.addLayout(history_actions)

        refresh_btn = QPushButton("Refresh Purchases")
        refresh_btn.setObjectName("SecondaryPurchaseButton")
        refresh_btn.clicked.connect(self.refresh_purchases_tab)

        export_purchases_btn = QPushButton("Export CSV")
        export_purchases_btn.setObjectName("SecondaryPurchaseButton")
        export_purchases_btn.clicked.connect(self.export_purchases_csv)

        footer_row = QHBoxLayout()
        footer_row.addWidget(export_purchases_btn)
        footer_row.addWidget(refresh_btn)
        footer_row.addStretch()

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(6)
        left_layout.addWidget(self.purchase_lines_table)
        left_layout.addLayout(table_actions_row)
        left_layout.addWidget(self.purchase_empty_state_label)
        left_layout.addWidget(total_panel)
        left_layout.addWidget(purchase_hint)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(6)
        right_layout.addWidget(history_box)
        right_layout.addLayout(footer_row)

        workspace_splitter = QSplitter(Qt.Horizontal)
        workspace_splitter.setChildrenCollapsible(False)
        workspace_splitter.addWidget(left_panel)
        workspace_splitter.addWidget(right_panel)
        workspace_splitter.setStretchFactor(0, 3)
        workspace_splitter.setStretchFactor(1, 2)

        root_layout.addWidget(entry_box)
        root_layout.addWidget(purchase_builder_hint)
        root_layout.addWidget(self.purchase_mode_label)
        root_layout.addWidget(self.purchase_feedback_label)
        root_layout.addWidget(workspace_splitter)
        root_layout.setStretch(4, 1)

        self.purchase_qty_spin.lineEdit().returnPressed.connect(self.add_purchase_line)
        self.purchase_cost_spin.lineEdit().returnPressed.connect(self.add_purchase_line)
        self._apply_purchases_panel_style(tab)
        self._update_purchase_stock_preview()
        self._set_purchase_filter_button_state("custom")
        return tab

    def _build_expenses_tab(self) -> QWidget:
        tab = QWidget()
        tab.setObjectName("ExpensesPanel")
        root_layout = QVBoxLayout(tab)

        entry_box = QGroupBox("Record Expense")
        form_layout = QGridLayout(entry_box)

        self.expense_type_combo = QComboBox()
        self.expense_type_combo.setEditable(True)
        self.expense_type_combo.addItems(["Rent", "Electricity", "Wages", "Misc"])

        self.expense_amount_spin = QDoubleSpinBox()
        self.expense_amount_spin.setDecimals(2)
        self.expense_amount_spin.setMinimum(0.01)
        self.expense_amount_spin.setMaximum(10000000)
        self.expense_amount_spin.setPrefix("INR ")

        self.expense_notes_input = QLineEdit()
        self.expense_notes_input.setPlaceholderText("Optional notes")

        add_expense_btn = QPushButton("Add Expense")
        add_expense_btn.setObjectName("PrimaryExpenseButton")
        add_expense_btn.clicked.connect(self.add_expense)

        form_layout.addWidget(QLabel("<b>Type</b>"), 0, 0)
        form_layout.addWidget(self.expense_type_combo, 0, 1)
        form_layout.addWidget(QLabel("<b>Amount</b>"), 0, 2)
        form_layout.addWidget(self.expense_amount_spin, 0, 3)
        form_layout.addWidget(QLabel("<b>Notes</b>"), 1, 0)
        form_layout.addWidget(self.expense_notes_input, 1, 1, 1, 3)
        form_layout.addWidget(add_expense_btn, 2, 3)

        history_box = QGroupBox("Recent Expenses")
        history_layout = QVBoxLayout(history_box)
        self.expense_history_table = QTableWidget(0, 5)
        self.expense_history_table.setHorizontalHeaderLabels(["ID", "Type", "Amount", "Date", "Notes"])
        self.expense_history_table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.SelectedClicked | QTableWidget.EditKeyPressed
        )
        self.expense_history_table.setAlternatingRowColors(True)
        self.expense_history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.expense_history_table.verticalHeader().setVisible(False)
        self.expense_history_table.verticalHeader().setDefaultSectionSize(34)
        self.expense_history_table.itemChanged.connect(self._on_expense_item_changed)
        self.expense_history_table.horizontalHeader().setStretchLastSection(True)
        history_layout.addWidget(self.expense_history_table)

        expense_hint = QLabel("Tip: Double-click Type, Amount, or Notes in Recent Expenses to edit inline.")

        filter_box = QGroupBox("Expense Date Filter")
        filter_layout = QHBoxLayout(filter_box)
        self.expense_from_date = QDateEdit()
        self.expense_from_date.setCalendarPopup(True)
        self.expense_from_date.setDisplayFormat("yyyy-MM-dd")
        self.expense_from_date.setDate(QDate.currentDate())

        self.expense_to_date = QDateEdit()
        self.expense_to_date.setCalendarPopup(True)
        self.expense_to_date.setDisplayFormat("yyyy-MM-dd")
        self.expense_to_date.setDate(QDate.currentDate())

        e_today_btn = QPushButton("Today")
        e_today_btn.setObjectName("ExpenseFilterPreset")
        e_today_btn.clicked.connect(
            lambda: self._apply_quick_range(
                self.expense_from_date,
                self.expense_to_date,
                "today",
                self.refresh_expenses_tab,
            )
        )
        e_7d_btn = QPushButton("Last 7 Days")
        e_7d_btn.setObjectName("ExpenseFilterPreset")
        e_7d_btn.clicked.connect(
            lambda: self._apply_quick_range(
                self.expense_from_date,
                self.expense_to_date,
                "last7",
                self.refresh_expenses_tab,
            )
        )
        e_month_btn = QPushButton("This Month")
        e_month_btn.setObjectName("ExpenseFilterPreset")
        e_month_btn.clicked.connect(
            lambda: self._apply_quick_range(
                self.expense_from_date,
                self.expense_to_date,
                "month",
                self.refresh_expenses_tab,
            )
        )
        e_custom_btn = QPushButton("Custom")
        e_custom_btn.setObjectName("ExpenseFilterPreset")
        e_custom_btn.clicked.connect(self.refresh_expenses_tab)
        e_apply_btn = QPushButton("Apply")
        e_apply_btn.setObjectName("ExpenseFilterPreset")
        e_apply_btn.clicked.connect(self.refresh_expenses_tab)

        filter_layout.addWidget(QLabel("<b>From</b>"))
        filter_layout.addWidget(self.expense_from_date)
        filter_layout.addWidget(QLabel("<b>To</b>"))
        filter_layout.addWidget(self.expense_to_date)
        filter_layout.addWidget(e_today_btn)
        filter_layout.addWidget(e_7d_btn)
        filter_layout.addWidget(e_month_btn)
        filter_layout.addWidget(e_custom_btn)
        filter_layout.addWidget(e_apply_btn)
        filter_layout.addStretch()

        refresh_btn = QPushButton("Refresh Expenses")
        refresh_btn.setObjectName("SecondaryExpenseButton")
        refresh_btn.clicked.connect(self.refresh_expenses_tab)

        export_expenses_btn = QPushButton("Export CSV")
        export_expenses_btn.setObjectName("SecondaryExpenseButton")
        export_expenses_btn.clicked.connect(self.export_expenses_csv)

        root_layout.addWidget(entry_box)
        root_layout.addWidget(expense_hint)
        root_layout.addWidget(filter_box)
        root_layout.addWidget(history_box)
        root_layout.addWidget(export_expenses_btn)
        root_layout.addWidget(refresh_btn)
        self._apply_expenses_panel_style(tab)
        return tab

    def _apply_expenses_panel_style(self, tab: QWidget) -> None:
        tab.setStyleSheet(
            "#ExpensesPanel QGroupBox {"
            "border: 1px solid #3f485b;"
            "border-radius: 8px;"
            "margin-top: 8px;"
            "padding-top: 8px;"
            "}"
            "#ExpensesPanel QGroupBox::title {"
            "subcontrol-origin: margin;"
            "left: 8px;"
            "padding: 0 4px;"
            "font-weight: 600;"
            "color: #dbe5f5;"
            "}"
            "#ExpensesPanel QTableWidget {"
            "gridline-color: #2f4261;"
            "background-color: #152036;"
            "alternate-background-color: #1b2942;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #385176;"
            "border-radius: 6px;"
            "}"
            "#ExpensesPanel QTableWidget::item:hover {"
            "background-color: #203a5f;"
            "}"
            "#ExpensesPanel QHeaderView::section {"
            "background-color: #233552;"
            "color: #eaf1ff;"
            "padding: 7px 8px;"
            "border: 0px;"
            "border-right: 1px solid #3f5980;"
            "font-weight: 600;"
            "}"
            "#ExpensesPanel QLineEdit,"
            "#ExpensesPanel QDoubleSpinBox,"
            "#ExpensesPanel QComboBox,"
            "#ExpensesPanel QDateEdit {"
            "background-color: #131d2e;"
            "color: #f2f6ff;"
            "border: 1px solid #425a80;"
            "border-radius: 6px;"
            "padding: 6px 8px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "#ExpensesPanel QLineEdit::placeholder {"
            "color: #90a4c8;"
            "}"
            "#ExpensesPanel QComboBox::drop-down,"
            "#ExpensesPanel QDateEdit::drop-down {"
            "border: 0px;"
            "width: 22px;"
            "}"
            "#ExpensesPanel QComboBox QAbstractItemView,"
            "#ExpensesPanel QDateEdit QAbstractItemView {"
            "background-color: #1a263c;"
            "color: #edf3ff;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "border: 1px solid #435e88;"
            "}"
            "#ExpensesPanel QLineEdit:focus,"
            "#ExpensesPanel QDoubleSpinBox:focus,"
            "#ExpensesPanel QComboBox:focus,"
            "#ExpensesPanel QDateEdit:focus {"
            "border: 1px solid #55c2ff;"
            "background-color: #1a2940;"
            "}"
            "#ExpensesPanel QPushButton {"
            "border-radius: 8px;"
            "padding: 7px 12px;"
            "border: 1px solid #4a5872;"
            "background-color: #2d3a50;"
            "color: #edf3fb;"
            "}"
            "#ExpensesPanel QPushButton#PrimaryExpenseButton {"
            "background-color: #219552;"
            "border: 1px solid #25a55a;"
            "font-weight: 700;"
            "}"
            "#ExpensesPanel QPushButton#PrimaryExpenseButton:hover {"
            "background-color: #2bb061;"
            "}"
            "#ExpensesPanel QPushButton#SecondaryExpenseButton {"
            "background-color: #2d3a50;"
            "}"
            "#ExpensesPanel QPushButton#ExpenseFilterPreset {"
            "background-color: #334767;"
            "border-color: #556e95;"
            "}"
            "#ExpensesPanel QPushButton#ExpenseFilterPreset:hover {"
            "background-color: #3e5780;"
            "}"
        )

    def _build_reports_tab(self) -> QWidget:
        tab = QWidget()
        tab.setObjectName("ReportsPanel")
        root_layout = QVBoxLayout(tab)

        title = QLabel("Operations Dashboard")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #e0e0e0;")

        subtitle = QLabel("Live performance snapshot for sales, stock, and profitability")
        subtitle.setStyleSheet("color: #9ea7b8;")

        cards_grid = QGridLayout()
        cards_grid.setHorizontalSpacing(10)
        cards_grid.setVerticalSpacing(10)

        sales_card, self.sales_value = self._make_report_metric_card("Sales", "#4caf50")
        cogs_card, self.cogs_value = self._make_report_metric_card("COGS", "#ffb74d")
        gross_card, self.gross_profit_value = self._make_report_metric_card("Gross Profit", "#64b5f6")
        purchases_card, self.purchases_value = self._make_report_metric_card("Purchases", "#9575cd")
        expenses_card, self.expenses_value = self._make_report_metric_card("Expenses", "#ef5350")
        fixed_daily_card, self.fixed_daily_value = self._make_report_metric_card("Fixed Cost / Day", "#ff7043")
        net_card, self.net_value = self._make_report_metric_card("Realistic Daily Profit", "#26c6da")

        metric_cards = [
            sales_card,
            cogs_card,
            gross_card,
            purchases_card,
            expenses_card,
            fixed_daily_card,
            net_card,
        ]
        for index, card in enumerate(metric_cards):
            row = index // 4
            col = index % 4
            cards_grid.addWidget(card, row, col)
        for col in range(4):
            cards_grid.setColumnStretch(col, 1)

        fixed_cost_box = QGroupBox("Monthly Fixed Costs (Auto-divided daily)")
        fixed_cost_layout = QGridLayout(fixed_cost_box)

        self.rent_spin = QDoubleSpinBox()
        self.rent_spin.setPrefix("INR ")
        self.rent_spin.setMaximum(100000000)

        self.salary_spin = QDoubleSpinBox()
        self.salary_spin.setPrefix("INR ")
        self.salary_spin.setMaximum(100000000)

        self.maintenance_spin = QDoubleSpinBox()
        self.maintenance_spin.setPrefix("INR ")
        self.maintenance_spin.setMaximum(100000000)

        self.electricity_spin = QDoubleSpinBox()
        self.electricity_spin.setPrefix("INR ")
        self.electricity_spin.setMaximum(100000000)

        self.monthly_fixed_total_label = QLabel("Monthly Fixed Total: INR 0.00")
        self.monthly_fixed_total_label.setStyleSheet("font-weight: bold;")

        save_fixed_btn = QPushButton("Save Fixed Costs")
        save_fixed_btn.setObjectName("DataOpsPrimaryButton")
        save_fixed_btn.clicked.connect(self.save_monthly_fixed_costs)

        fixed_cost_layout.addWidget(QLabel("Rent"), 0, 0)
        fixed_cost_layout.addWidget(self.rent_spin, 0, 1)
        fixed_cost_layout.addWidget(QLabel("Salaries"), 0, 2)
        fixed_cost_layout.addWidget(self.salary_spin, 0, 3)

        fixed_cost_layout.addWidget(QLabel("Maintenance"), 1, 0)
        fixed_cost_layout.addWidget(self.maintenance_spin, 1, 1)
        fixed_cost_layout.addWidget(QLabel("Electricity"), 1, 2)
        fixed_cost_layout.addWidget(self.electricity_spin, 1, 3)

        fixed_cost_layout.addWidget(self.monthly_fixed_total_label, 2, 0, 1, 3)
        fixed_cost_layout.addWidget(save_fixed_btn, 2, 3)

        daily_overhead_box = QGroupBox("Daily Variable Overhead (for Recipe Costing)")
        daily_overhead_layout = QGridLayout(daily_overhead_box)
        self.overhead_date_edit = QDateEdit()
        self.overhead_date_edit.setCalendarPopup(True)
        self.overhead_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.overhead_date_edit.setDate(QDate.currentDate())
        self.overhead_date_edit.dateChanged.connect(self.refresh_reports)

        self.overhead_gas_spin = QDoubleSpinBox()
        self.overhead_gas_spin.setPrefix("INR ")
        self.overhead_gas_spin.setMaximum(100000000)
        self.overhead_labor_spin = QDoubleSpinBox()
        self.overhead_labor_spin.setPrefix("INR ")
        self.overhead_labor_spin.setMaximum(100000000)
        self.overhead_misc_spin = QDoubleSpinBox()
        self.overhead_misc_spin.setPrefix("INR ")
        self.overhead_misc_spin.setMaximum(100000000)
        self.overhead_units_spin = QDoubleSpinBox()
        self.overhead_units_spin.setMaximum(100000000)

        save_overhead_btn = QPushButton("Save Daily Overhead")
        save_overhead_btn.setObjectName("DataOpsPrimaryButton")
        save_overhead_btn.clicked.connect(self.save_daily_overhead)

        self.overhead_per_unit_label = QLabel("Overhead / Unit: INR 0.00")
        self.overhead_per_unit_label.setStyleSheet("font-weight: bold;")

        daily_overhead_layout.addWidget(QLabel("Date"), 0, 0)
        daily_overhead_layout.addWidget(self.overhead_date_edit, 0, 1)
        daily_overhead_layout.addWidget(QLabel("Gas"), 0, 2)
        daily_overhead_layout.addWidget(self.overhead_gas_spin, 0, 3)
        daily_overhead_layout.addWidget(QLabel("Labor"), 1, 0)
        daily_overhead_layout.addWidget(self.overhead_labor_spin, 1, 1)
        daily_overhead_layout.addWidget(QLabel("Misc"), 1, 2)
        daily_overhead_layout.addWidget(self.overhead_misc_spin, 1, 3)
        daily_overhead_layout.addWidget(QLabel("Expected Units"), 2, 0)
        daily_overhead_layout.addWidget(self.overhead_units_spin, 2, 1)
        daily_overhead_layout.addWidget(self.overhead_per_unit_label, 2, 2)
        daily_overhead_layout.addWidget(save_overhead_btn, 2, 3)

        backup_automation_box = QGroupBox("Backup Automation")
        backup_auto_layout = QHBoxLayout(backup_automation_box)
        self.auto_backup_enabled_checkbox = QCheckBox("Enable scheduled backup")
        self.auto_backup_interval_spin = QSpinBox()
        self.auto_backup_interval_spin.setRange(5, 1440)
        self.auto_backup_interval_spin.setSuffix(" min")
        auto_backup_save_btn = QPushButton("Save Backup Schedule")
        auto_backup_save_btn.setObjectName("DataOpsPrimaryButton")
        auto_backup_save_btn.clicked.connect(self.save_backup_preferences)
        backup_auto_layout.addWidget(self.auto_backup_enabled_checkbox)
        backup_auto_layout.addWidget(QLabel("Interval"))
        backup_auto_layout.addWidget(self.auto_backup_interval_spin)
        backup_auto_layout.addWidget(auto_backup_save_btn)
        backup_auto_layout.addStretch()

        trend_box = QGroupBox("Sales Trend (Selected Range)")
        trend_layout = QVBoxLayout(trend_box)
        self.sales_trend_table = QTableWidget(0, 5)
        self.sales_trend_table.setHorizontalHeaderLabels(
            ["Date", "Bills", "Sales", "COGS", "Gross Profit"]
        )
        self._style_report_table(self.sales_trend_table)
        self._apply_report_column_modes(
            self.sales_trend_table,
            [
                QHeaderView.Stretch,
                QHeaderView.Stretch,
                QHeaderView.Stretch,
                QHeaderView.Stretch,
                QHeaderView.Stretch,
            ],
        )
        self.sales_trend_table.setMinimumHeight(240)
        trend_layout.addWidget(self.sales_trend_table)

        top_items_box = QGroupBox("Top Selling Items")
        top_items_layout = QVBoxLayout(top_items_box)
        self.top_items_table = QTableWidget(0, 3)
        self.top_items_table.setHorizontalHeaderLabels(["Item", "Qty Sold", "Sales Value"])
        self._style_report_table(self.top_items_table)
        self._apply_report_column_modes(
            self.top_items_table,
            [
                QHeaderView.Interactive,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
            ],
        )
        self.top_items_table.setMinimumHeight(240)
        self.top_items_table.setColumnWidth(0, 220)
        top_items_layout.addWidget(self.top_items_table)

        low_stock_box = QGroupBox("Low Stock")
        low_stock_layout = QVBoxLayout(low_stock_box)
        self.low_stock_table = QTableWidget(0, 3)
        self.low_stock_table.setHorizontalHeaderLabels(["Item", "Stock", "Reorder"])
        self._style_report_table(self.low_stock_table)
        self._apply_report_column_modes(
            self.low_stock_table,
            [
                QHeaderView.Interactive,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
            ],
        )
        self.low_stock_table.setColumnWidth(0, 220)
        low_stock_layout.addWidget(self.low_stock_table)

        ledger_box = QGroupBox("Stock Movement Ledger")
        ledger_layout = QVBoxLayout(ledger_box)
        self.ledger_table = QTableWidget(0, 6)
        self.ledger_table.setHorizontalHeaderLabels(
            ["Time", "Item", "Change", "Reason", "Reference", "Notes"]
        )
        self._style_report_table(self.ledger_table)
        self._apply_report_column_modes(
            self.ledger_table,
            [
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.Interactive,
            ],
        )
        self.ledger_table.setColumnWidth(5, 320)
        ledger_layout.addWidget(self.ledger_table)

        refresh_reports_btn = QPushButton("Refresh Reports")
        refresh_reports_btn.setObjectName("PrimaryReportButton")
        refresh_reports_btn.clicked.connect(self.refresh_reports)

        close_day_btn = QPushButton("Close Day (Ctrl+L)")
        close_day_btn.setObjectName("DataOpsWarningButton")
        close_day_btn.clicked.connect(self.close_day)

        backup_btn = QPushButton("Backup Now")
        backup_btn.setObjectName("DataOpsPrimaryButton")
        backup_btn.clicked.connect(self.backup_now)

        export_btn = QPushButton("Export DB Backup")
        export_btn.setObjectName("DataOpsExportButton")
        export_btn.clicked.connect(self.export_backup_dialog)

        restore_btn = QPushButton("Restore DB Backup")
        restore_btn.setObjectName("DataOpsDangerButton")
        restore_btn.clicked.connect(self.restore_backup_dialog)

        export_reports_btn = QPushButton("Export CSV")
        export_reports_btn.setObjectName("DataOpsExportButton")
        export_reports_btn.clicked.connect(self.export_reports_csv)

        export_reports_xlsx_btn = QPushButton("Export XLSX")
        export_reports_xlsx_btn.setObjectName("DataOpsExportButton")
        export_reports_xlsx_btn.clicked.connect(self.export_reports_xlsx)

        export_all_btn = QPushButton("Export All CSV")
        export_all_btn.setObjectName("DataOpsExportButton")
        export_all_btn.clicked.connect(self.export_all_csv)

        print_summary_btn = QPushButton("Print Summary")
        print_summary_btn.setObjectName("DataOpsExportButton")
        print_summary_btn.clicked.connect(self.export_printable_summary)

        self.report_from_date = QDateEdit()
        self.report_from_date.setCalendarPopup(True)
        self.report_from_date.setDisplayFormat("yyyy-MM-dd")
        self.report_from_date.setDate(QDate.currentDate().addDays(-6))
        self.report_from_date.setMinimumWidth(132)

        self.report_to_date = QDateEdit()
        self.report_to_date.setCalendarPopup(True)
        self.report_to_date.setDisplayFormat("yyyy-MM-dd")
        self.report_to_date.setDate(QDate.currentDate())
        self.report_to_date.setMinimumWidth(132)

        r_today_btn = QPushButton("Today")
        r_today_btn.setObjectName("ReportFilterPreset")
        r_today_btn.clicked.connect(
            lambda: self._apply_quick_range(
                self.report_from_date,
                self.report_to_date,
                "today",
                self.refresh_reports,
            )
        )
        r_7d_btn = QPushButton("Last 7 Days")
        r_7d_btn.setObjectName("ReportFilterPreset")
        r_7d_btn.clicked.connect(
            lambda: self._apply_quick_range(
                self.report_from_date,
                self.report_to_date,
                "last7",
                self.refresh_reports,
            )
        )
        r_month_btn = QPushButton("This Month")
        r_month_btn.setObjectName("ReportFilterPreset")
        r_month_btn.clicked.connect(
            lambda: self._apply_quick_range(
                self.report_from_date,
                self.report_to_date,
                "month",
                self.refresh_reports,
            )
        )
        r_custom_btn = QPushButton("Custom")
        r_custom_btn.setObjectName("ReportFilterPreset")
        r_custom_btn.clicked.connect(self.refresh_reports)

        self.top_items_limit_spin = QSpinBox()
        self.top_items_limit_spin.setRange(5, 100)
        self.top_items_limit_spin.setValue(20)
        self.ledger_limit_spin = QSpinBox()
        self.ledger_limit_spin.setRange(50, 2000)
        self.ledger_limit_spin.setValue(500)

        self.role_combo = QComboBox()
        self.role_combo.addItems(["cashier", "admin"])
        self.role_combo.setCurrentText(self.current_role)
        self.role_combo.currentTextChanged.connect(self.on_role_changed)

        self.open_after_export_checkbox = QCheckBox("Open after export")
        self.open_after_export_checkbox.setChecked(True)
        reports_tabs = QTabWidget()

        overview_tab = QWidget()
        overview_tab.setObjectName("ReportsOverviewTab")
        overview_layout = QVBoxLayout(overview_tab)
        overview_filters_box = QGroupBox("Overview Filters")
        overview_filters_layout = QGridLayout(overview_filters_box)
        overview_filters_layout.setHorizontalSpacing(10)
        overview_filters_layout.setVerticalSpacing(8)
        overview_filters_layout.addWidget(QLabel("<b>From</b>"), 0, 0)
        overview_filters_layout.addWidget(self.report_from_date, 0, 1)
        overview_filters_layout.addWidget(QLabel("<b>To</b>"), 0, 2)
        overview_filters_layout.addWidget(self.report_to_date, 0, 3)
        overview_filters_layout.addWidget(r_today_btn, 0, 4)
        overview_filters_layout.addWidget(r_7d_btn, 0, 5)
        overview_filters_layout.addWidget(r_month_btn, 0, 6)
        overview_filters_layout.addWidget(r_custom_btn, 0, 7)
        overview_filters_layout.addWidget(QLabel("<b>Top Items</b>"), 1, 0)
        overview_filters_layout.addWidget(self.top_items_limit_spin, 1, 1)
        overview_filters_layout.addWidget(refresh_reports_btn, 1, 2)
        overview_filters_layout.setColumnStretch(8, 1)

        overview_tables_splitter = QSplitter(Qt.Horizontal)
        overview_tables_splitter.setChildrenCollapsible(False)
        overview_tables_splitter.addWidget(trend_box)
        overview_tables_splitter.addWidget(top_items_box)
        overview_tables_splitter.setStretchFactor(0, 3)
        overview_tables_splitter.setStretchFactor(1, 2)

        overview_layout.addWidget(title)
        overview_layout.addWidget(subtitle)
        overview_layout.addWidget(overview_filters_box)
        overview_layout.addLayout(cards_grid)
        overview_layout.addWidget(overview_tables_splitter)

        stock_tab = QWidget()
        stock_layout = QVBoxLayout(stock_tab)
        stock_refresh_btn = QPushButton("Refresh Stock and Audit")
        stock_refresh_btn.clicked.connect(self.refresh_reports)
        stock_controls_box = QGroupBox("Stock Analysis Controls")
        stock_controls_layout = QHBoxLayout(stock_controls_box)
        stock_controls_layout.addWidget(QLabel("Ledger"))
        stock_controls_layout.addWidget(self.ledger_limit_spin)
        stock_controls_layout.addWidget(stock_refresh_btn)
        stock_controls_layout.addStretch()

        stock_layout.addWidget(stock_controls_box)
        stock_layout.addWidget(low_stock_box)
        stock_layout.addWidget(ledger_box)

        audit_box = QGroupBox("Audit Log")
        audit_box_layout = QVBoxLayout(audit_box)
        self.audit_log_table = QTableWidget(0, 6)
        self.audit_log_table.setHorizontalHeaderLabels(["Time", "Role", "Action", "Entity", "Entity ID", "Details"])
        self._style_report_table(self.audit_log_table)
        self._apply_report_column_modes(
            self.audit_log_table,
            [
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.Interactive,
            ],
        )
        self.audit_log_table.setColumnWidth(5, 340)
        refresh_audit_btn = QPushButton("Refresh Audit")
        refresh_audit_btn.clicked.connect(self.refresh_reports)
        export_audit_csv_btn = QPushButton("Export Audit CSV")
        export_audit_csv_btn.clicked.connect(self.export_audit_csv)
        export_audit_xlsx_btn = QPushButton("Export Audit XLSX")
        export_audit_xlsx_btn.clicked.connect(self.export_audit_xlsx)
        audit_actions = QHBoxLayout()
        audit_actions.addWidget(refresh_audit_btn)
        audit_actions.addWidget(export_audit_csv_btn)
        audit_actions.addWidget(export_audit_xlsx_btn)
        audit_actions.addStretch()
        audit_box_layout.addWidget(self.audit_log_table)
        audit_box_layout.addLayout(audit_actions)

        costing_exceptions_box = QGroupBox("Costing Exceptions")
        costing_exceptions_layout = QVBoxLayout(costing_exceptions_box)
        self.costing_exceptions_table = QTableWidget(0, 6)
        self.costing_exceptions_table.setHorizontalHeaderLabels(
            ["Time", "Type", "Item", "Sale ID", "Item ID", "Details"]
        )
        self._style_report_table(self.costing_exceptions_table)
        self._apply_report_column_modes(
            self.costing_exceptions_table,
            [
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.ResizeToContents,
                QHeaderView.Stretch,
            ],
        )
        costing_exceptions_layout.addWidget(self.costing_exceptions_table)

        stock_layout.addWidget(audit_box)
        stock_layout.addWidget(costing_exceptions_box)

        data_ops_tab = QWidget()
        data_ops_tab.setObjectName("ReportsDataOpsTab")
        data_ops_layout = QVBoxLayout(data_ops_tab)
        data_ops_layout.setSpacing(8)

        data_ops_buttons = [
            close_day_btn,
            backup_btn,
            export_btn,
            restore_btn,
            export_reports_btn,
            export_reports_xlsx_btn,
            export_all_btn,
            print_summary_btn,
            auto_backup_save_btn,
            save_fixed_btn,
        ]
        for button in data_ops_buttons:
            button.setMinimumHeight(34)
            button.setMaximumHeight(34)

        section_style = (
            "QGroupBox {"
            "border: 1px solid #3a5780;"
            "border-radius: 6px;"
            "margin-top: 10px;"
            "padding-top: 8px;"
            "background-color: #1b2a40;"
            "}"
            "QGroupBox::title {"
            "subcontrol-origin: margin;"
            "left: 10px;"
            "padding: 0 4px;"
            "color: #e5efff;"
            "font-weight: 600;"
            "}"
        )

        role_and_close_box = QGroupBox("Access and Closing")
        role_and_close_box.setObjectName("DataOpsSectionBox")
        role_and_close_box.setStyleSheet(section_style)
        role_and_close_layout = QHBoxLayout(role_and_close_box)
        role_and_close_layout.addWidget(QLabel("Role"))
        role_and_close_layout.addWidget(self.role_combo)
        role_and_close_layout.addWidget(close_day_btn)
        role_and_close_layout.addStretch()

        backup_actions_box = QGroupBox("Backup and Restore")
        backup_actions_box.setObjectName("DataOpsSectionBox")
        backup_actions_box.setStyleSheet(section_style)
        backup_actions_layout = QHBoxLayout(backup_actions_box)
        backup_actions_layout.addWidget(backup_btn)
        backup_actions_layout.addWidget(export_btn)
        backup_actions_layout.addWidget(restore_btn)
        backup_actions_layout.addStretch()

        report_exports_box = QGroupBox("Reporting Exports")
        report_exports_box.setObjectName("DataOpsSectionBox")
        report_exports_box.setStyleSheet(section_style)
        report_exports_layout = QHBoxLayout(report_exports_box)
        report_exports_layout.addWidget(export_reports_btn)
        report_exports_layout.addWidget(export_reports_xlsx_btn)
        report_exports_layout.addWidget(export_all_btn)
        report_exports_layout.addWidget(print_summary_btn)
        report_exports_layout.addWidget(self.open_after_export_checkbox)
        report_exports_layout.addStretch()

        backup_automation_box.setObjectName("DataOpsSectionBox")
        fixed_cost_box.setObjectName("DataOpsSectionBox")
        daily_overhead_box.setObjectName("DataOpsSectionBox")
        backup_automation_box.setStyleSheet(section_style)
        fixed_cost_box.setStyleSheet(section_style)
        daily_overhead_box.setStyleSheet(section_style)

        separator_one = QFrame()
        separator_one.setFrameShape(QFrame.HLine)
        separator_one.setStyleSheet("color: #2b4363;")

        separator_two = QFrame()
        separator_two.setFrameShape(QFrame.HLine)
        separator_two.setStyleSheet("color: #2b4363;")

        separator_three = QFrame()
        separator_three.setFrameShape(QFrame.HLine)
        separator_three.setStyleSheet("color: #2b4363;")

        separator_four = QFrame()
        separator_four.setFrameShape(QFrame.HLine)
        separator_four.setStyleSheet("color: #2b4363;")

        separator_five = QFrame()
        separator_five.setFrameShape(QFrame.HLine)
        separator_five.setStyleSheet("color: #2b4363;")

        data_ops_layout.addWidget(role_and_close_box)
        data_ops_layout.addWidget(separator_one)
        data_ops_layout.addWidget(backup_actions_box)
        data_ops_layout.addWidget(separator_two)
        data_ops_layout.addWidget(backup_automation_box)
        data_ops_layout.addWidget(separator_three)
        data_ops_layout.addWidget(fixed_cost_box)
        data_ops_layout.addWidget(separator_four)
        data_ops_layout.addWidget(daily_overhead_box)
        data_ops_layout.addWidget(separator_five)
        data_ops_layout.addWidget(report_exports_box)
        data_ops_layout.addStretch()

        reports_tabs.addTab(overview_tab, "Overview")
        reports_tabs.addTab(stock_tab, "Stock and Audit")
        reports_tabs.addTab(data_ops_tab, "Data Ops")

        reports_content = QWidget()
        reports_content.setObjectName("ReportsContent")
        reports_content_layout = QVBoxLayout(reports_content)
        reports_content_layout.setContentsMargins(8, 8, 8, 8)
        reports_content_layout.setSpacing(10)
        reports_content_layout.addWidget(reports_tabs)

        reports_scroll = QScrollArea()
        reports_scroll.setWidgetResizable(True)
        reports_scroll.setWidget(reports_content)

        root_layout.addWidget(reports_scroll)
        self._apply_report_table_width_profiles()
        self._apply_reports_panel_style(tab)

        return tab

    def _apply_reports_panel_style(self, tab: QWidget) -> None:
        tab.setStyleSheet(
            "#ReportsPanel {"
            "background-color: #141d2d;"
            "}"
            "#ReportsContent, #ReportsOverviewTab {"
            "background-color: #141d2d;"
            "}"
            "#ReportsDataOpsTab {"
            "background-color: #141d2d;"
            "}"
            "#ReportsDataOpsTab QGroupBox#DataOpsSectionBox {"
            "background-color: #1b2a40;"
            "border: 1px solid #3a5780;"
            "}"
            "#ReportsPanel QScrollArea {"
            "background-color: #141d2d;"
            "border: 0px;"
            "}"
            "#ReportsPanel QScrollArea > QWidget > QWidget {"
            "background-color: #141d2d;"
            "}"
            "#ReportsPanel QGroupBox {"
            "border: 1px solid #3f4b61;"
            "border-radius: 8px;"
            "margin-top: 8px;"
            "padding-top: 8px;"
            "background-color: #202734;"
            "}"
            "#ReportsPanel QGroupBox::title {"
            "subcontrol-origin: margin;"
            "left: 8px;"
            "padding: 0 4px;"
            "font-weight: 600;"
            "color: #dce6f7;"
            "}"
            "#ReportsPanel QDateEdit,"
            "#ReportsPanel QSpinBox,"
            "#ReportsPanel QDoubleSpinBox,"
            "#ReportsPanel QComboBox {"
            "background-color: #121c2d;"
            "color: #f1f6ff;"
            "border: 1px solid #435b80;"
            "border-radius: 6px;"
            "padding: 5px 8px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "#ReportsPanel QDateEdit::drop-down,"
            "#ReportsPanel QComboBox::drop-down,"
            "#ReportsPanel QSpinBox::up-button,"
            "#ReportsPanel QSpinBox::down-button {"
            "border: 0px;"
            "width: 18px;"
            "}"
            "#ReportsPanel QDateEdit:focus,"
            "#ReportsPanel QSpinBox:focus,"
            "#ReportsPanel QDoubleSpinBox:focus,"
            "#ReportsPanel QComboBox:focus {"
            "border: 1px solid #58c6ff;"
            "background-color: #1a2940;"
            "}"
            "#ReportsPanel QPushButton {"
            "border-radius: 8px;"
            "padding: 6px 12px;"
            "border: 1px solid #4a5872;"
            "background-color: #2d3a50;"
            "color: #edf3fb;"
            "}"
            "#ReportsPanel QPushButton#PrimaryReportButton {"
            "background-color: #2f7ee8;"
            "border-color: #4e97f0;"
            "font-weight: 700;"
            "}"
            "#ReportsPanel QPushButton#PrimaryReportButton:hover {"
            "background-color: #3f8ff1;"
            "}"
            "#ReportsPanel QPushButton#ReportFilterPreset {"
            "background-color: #334767;"
            "border-color: #556e95;"
            "font-weight: 600;"
            "}"
            "#ReportsPanel QPushButton#ReportFilterPreset:hover {"
            "background-color: #3e5780;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsPrimaryButton {"
            "background-color: #1e9a58;"
            "border-color: #35b86f;"
            "color: #ffffff;"
            "font-weight: 700;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsPrimaryButton:hover {"
            "background-color: #27ac66;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsWarningButton {"
            "background-color: #a56a1f;"
            "border-color: #c88a3b;"
            "color: #fff7e8;"
            "font-weight: 700;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsWarningButton:hover {"
            "background-color: #b77b2d;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsDangerButton {"
            "background-color: #9b2f3a;"
            "border-color: #c04c5a;"
            "color: #ffffff;"
            "font-weight: 700;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsDangerButton:hover {"
            "background-color: #ad3a47;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsExportButton {"
            "background-color: #2f4f78;"
            "border-color: #4f74a5;"
            "color: #eaf3ff;"
            "font-weight: 600;"
            "}"
            "#ReportsDataOpsTab QPushButton#DataOpsExportButton:hover {"
            "background-color: #3b5f8f;"
            "}"
        )

    def _make_report_metric_card(self, heading: str, accent_color: str) -> tuple[QFrame, QLabel]:
        card = QFrame()
        card.setFrameShape(QFrame.StyledPanel)
        card.setStyleSheet(
            "QFrame {"
            "border: 1px solid #3a3f4b;"
            "border-radius: 8px;"
            "background-color: #1f232b;"
            "padding: 8px;"
            "}"
        )

        layout = QVBoxLayout(card)
        title = QLabel(heading)
        title.setStyleSheet(f"color: {accent_color}; font-weight: bold;")

        value = QLabel("INR 0.00")
        value.setStyleSheet("font-size: 17px; font-weight: bold; color: #f4f6fa;")

        layout.addWidget(title)
        layout.addWidget(value)
        layout.addStretch()
        return card, value

    def _style_report_table(self, table: QTableWidget) -> None:
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.verticalHeader().setVisible(False)
        table.setShowGrid(True)
        table.setWordWrap(False)
        table.setMinimumHeight(145)
        table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setMinimumSectionSize(72)
        table.horizontalHeader().setDefaultSectionSize(122)
        table.verticalHeader().setDefaultSectionSize(34)
        table.setStyleSheet(
            "QTableWidget {"
            "gridline-color: #304767;"
            "alternate-background-color: #192842;"
            "background-color: #142036;"
            "border: 1px solid #375377;"
            "border-radius: 6px;"
            "selection-background-color: #2b77e7;"
            "selection-color: #ffffff;"
            "}"
            "QTableWidget::item:hover {"
            "background-color: #1f3b61;"
            "}"
            "QHeaderView::section {"
            "background-color: #223654;"
            "color: #eaf1ff;"
            "padding: 8px 10px;"
            "border: 0px;"
            "border-right: 1px solid #3d5982;"
            "font-weight: 600;"
            "}"
        )

    def _apply_report_column_modes(self, table: QTableWidget, modes: list[QHeaderView.ResizeMode]) -> None:
        header = table.horizontalHeader()
        for idx, mode in enumerate(modes):
            header.setSectionResizeMode(idx, mode)

    def _apply_report_table_width_profiles(self) -> None:
        width = self.width()
        if width < 1200:
            item_col = 170
            ledger_notes_col = 260
            audit_details_col = 280
        elif width < 1500:
            item_col = 220
            ledger_notes_col = 320
            audit_details_col = 340
        else:
            item_col = 280
            ledger_notes_col = 420
            audit_details_col = 460

        self.top_items_table.setColumnWidth(0, item_col)
        self.low_stock_table.setColumnWidth(0, item_col)
        self.ledger_table.setColumnWidth(5, ledger_notes_col)
        self.audit_log_table.setColumnWidth(5, audit_details_col)

    @staticmethod
    def _report_item(value: str, right_align: bool = False) -> QTableWidgetItem:
        item = QTableWidgetItem(value)
        if right_align:
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
        else:
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        return item

    def _table_rows_for_csv(self, table: QTableWidget) -> list[list[str]]:
        headers = [table.horizontalHeaderItem(col).text() for col in range(table.columnCount())]
        rows: list[list[str]] = [headers]

        for row in range(table.rowCount()):
            values: list[str] = []
            for col in range(table.columnCount()):
                cell = table.item(row, col)
                values.append(cell.text() if cell is not None else "")
            rows.append(values)
        return rows

    @staticmethod
    def _set_date_edit_range(from_edit: QDateEdit, to_edit: QDateEdit, start: date, end: date) -> None:
        from_edit.setDate(QDate(start.year, start.month, start.day))
        to_edit.setDate(QDate(end.year, end.month, end.day))

    @staticmethod
    def _iso_range_from_edits(from_edit: QDateEdit, to_edit: QDateEdit) -> tuple[str, str]:
        start = from_edit.date().toPython()
        end = to_edit.date().toPython()
        if start > end:
            start, end = end, start
            from_edit.setDate(QDate(start.year, start.month, start.day))
            to_edit.setDate(QDate(end.year, end.month, end.day))
        return start.isoformat(), end.isoformat()

    def _apply_quick_range(
        self,
        from_edit: QDateEdit,
        to_edit: QDateEdit,
        preset: str,
        refresh_callback,
    ) -> None:
        today = date.today()
        if preset == "today":
            start, end = today, today
        elif preset == "last7":
            start, end = today - timedelta(days=6), today
        elif preset == "month":
            start, end = today.replace(day=1), today
        else:
            start, end = today, today

        self._set_date_edit_range(from_edit, to_edit, start, end)
        refresh_callback()

    @staticmethod
    def _range_suffix(start_date: str, end_date: str) -> str:
        if start_date == end_date:
            return start_date
        return f"{start_date}_to_{end_date}"

    def _save_csv_rows(self, default_name: str, rows: list[list[str]]) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save CSV",
            default_name,
            "CSV Files (*.csv)",
        )
        if not path:
            return

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle)
                writer.writerows(rows)
        except Exception as exc:
            QMessageBox.critical(self, "CSV Export", f"Failed to export CSV: {exc}")
            return

        try:
            if getattr(self, "open_after_export_checkbox", None) is None:
                should_open = True
            else:
                should_open = self.open_after_export_checkbox.isChecked()
            if should_open and hasattr(os, "startfile"):
                os.startfile(path)
        except Exception:
            pass

        QMessageBox.information(self, "CSV Export", f"CSV exported successfully:\n{path}")

    @staticmethod
    def _write_csv_file(path: str, rows: list[list[str]]) -> None:
        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerows(rows)

    def _build_reports_csv_rows(self, start_date: str, end_date: str) -> list[list[str]]:
        rows: list[list[str]] = []

        rows.append(["Report Range", start_date, end_date])
        rows.append([])
        rows.append(["Today Summary"])
        rows.append(["Sales", self.sales_value.text().replace("INR ", "")])
        rows.append(["COGS", self.cogs_value.text().replace("INR ", "")])
        rows.append(["Gross Profit", self.gross_profit_value.text().replace("INR ", "")])
        rows.append(["Purchases", self.purchases_value.text().replace("INR ", "")])
        rows.append(["Expenses", self.expenses_value.text().replace("INR ", "")])
        rows.append(["Fixed Cost / Day", self.fixed_daily_value.text().replace("INR ", "")])
        rows.append(["Realistic Daily Profit", self.net_value.text().replace("INR ", "")])
        rows.append(["Monthly Fixed Total", self.monthly_fixed_total_label.text().replace("Monthly Fixed Total: INR ", "")])
        rows.append([])

        rows.append(["Sales Trend"])
        rows.extend(self._table_rows_for_csv(self.sales_trend_table))
        rows.append([])

        rows.append(["Top Selling Items"])
        rows.extend(self._table_rows_for_csv(self.top_items_table))
        rows.append([])

        rows.append(["Low Stock"])
        rows.extend(self._table_rows_for_csv(self.low_stock_table))
        rows.append([])

        rows.append(["Stock Movement Ledger"])
        rows.extend(self._table_rows_for_csv(self.ledger_table))
        return rows

    def export_inventory_csv(self) -> None:
        rows = self._table_rows_for_csv(self.inventory_items_table)
        self._save_csv_rows("inventory_export.csv", rows)

    def export_purchases_csv(self) -> None:
        start_date, end_date = self._iso_range_from_edits(self.purchase_from_date, self.purchase_to_date)
        rows = self._table_rows_for_csv(self.purchase_history_table)
        self._save_csv_rows(f"purchases_{self._range_suffix(start_date, end_date)}.csv", rows)

    def export_expenses_csv(self) -> None:
        start_date, end_date = self._iso_range_from_edits(self.expense_from_date, self.expense_to_date)
        rows = self._table_rows_for_csv(self.expense_history_table)
        self._save_csv_rows(f"expenses_{self._range_suffix(start_date, end_date)}.csv", rows)

    def export_reports_csv(self) -> None:
        start_date, end_date = self._iso_range_from_edits(self.report_from_date, self.report_to_date)
        rows = self._build_reports_csv_rows(start_date=start_date, end_date=end_date)
        self._save_csv_rows(f"reports_{self._range_suffix(start_date, end_date)}.csv", rows)

    def export_all_csv(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Select Folder for CSV Export")
        if not folder:
            return

        report_start, report_end = self._iso_range_from_edits(self.report_from_date, self.report_to_date)

        files_to_export = {
            "inventory.csv": self._table_rows_for_csv(self.inventory_items_table),
            "purchases.csv": self._table_rows_for_csv(self.purchase_history_table),
            "expenses.csv": self._table_rows_for_csv(self.expense_history_table),
            "reports.csv": self._build_reports_csv_rows(start_date=report_start, end_date=report_end),
        }

        written_paths: list[str] = []
        try:
            for filename, rows in files_to_export.items():
                out_path = os.path.join(folder, filename)
                self._write_csv_file(out_path, rows)
                written_paths.append(out_path)
        except Exception as exc:
            QMessageBox.critical(self, "CSV Export", f"Failed during Export All: {exc}")
            return

        if self.open_after_export_checkbox.isChecked() and hasattr(os, "startfile"):
            try:
                os.startfile(folder)
            except Exception:
                pass

        QMessageBox.information(
            self,
            "CSV Export",
            "Export All completed:\n" + "\n".join(written_paths),
        )

    def save_monthly_fixed_costs(self) -> None:
        try:
            self.report_service.save_monthly_fixed_costs(
                rent=float(self.rent_spin.value()),
                salary=float(self.salary_spin.value()),
                maintenance=float(self.maintenance_spin.value()),
                electricity=float(self.electricity_spin.value()),
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Fixed Costs", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Fixed Costs Error", str(exc))
            return

        QMessageBox.information(self, "Fixed Costs", "Monthly fixed costs saved.")
        self.refresh_reports()

    def save_daily_overhead(self) -> None:
        pin = self._require_admin_access("Save Daily Overhead")
        if pin is None:
            return

        try:
            self.bookkeeping_service.set_daily_overhead(
                overhead_date=self.overhead_date_edit.date().toString("yyyy-MM-dd"),
                gas_cost=float(self.overhead_gas_spin.value()),
                labor_cost=float(self.overhead_labor_spin.value()),
                misc_cost=float(self.overhead_misc_spin.value()),
                expected_units=float(self.overhead_units_spin.value()),
                admin_pin=pin,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Daily Overhead", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Daily Overhead Error", str(exc))
            return

        QMessageBox.information(self, "Daily Overhead", "Daily overhead saved.")
        self.refresh_reports()

    def refresh_all(self) -> None:
        self.refresh_categories()
        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_purchases_tab()
        self.refresh_expenses_tab()
        self.refresh_reports()

    def refresh_categories(self) -> None:
        categories = self.inventory_service.list_categories()
        self.category_combo.clear()
        self.category_combo.addItem("Uncategorized", None)
        for category in categories:
            self.category_combo.addItem(category["name"], category["id"])

        if hasattr(self, "inventory_filter_category_combo"):
            current = self.inventory_filter_category_combo.currentData()
            self.inventory_filter_category_combo.blockSignals(True)
            self.inventory_filter_category_combo.clear()
            self.inventory_filter_category_combo.addItem("All Categories", None)
            for category in categories:
                self.inventory_filter_category_combo.addItem(category["name"], category["id"])
            index = self.inventory_filter_category_combo.findData(current)
            if index >= 0:
                self.inventory_filter_category_combo.setCurrentIndex(index)
            self.inventory_filter_category_combo.blockSignals(False)

    def refresh_inventory(self) -> None:
        self.inventory_items_cache = self.inventory_service.list_items()
        self.apply_inventory_filter()
        if hasattr(self, "inventory_inline_status_label"):
            self.inventory_inline_status_label.setVisible(False)

    def _on_inventory_item_changed(self, item: QTableWidgetItem) -> None:
        if self._updating_inventory_table:
            return

        column = item.column()
        if column not in (4, 6):
            return

        row = item.row()
        item_id = self._inventory_id_for_row(row)
        sell_cell = self.inventory_items_table.item(row, 4)
        reorder_cell = self.inventory_items_table.item(row, 6)
        if item_id is None or sell_cell is None or reorder_cell is None:
            return

        pin = self._require_admin_access("Inventory Edit")
        if pin is None:
            self.refresh_inventory()
            return

        try:
            selling_price = float(sell_cell.text().strip())
            reorder_level = float(reorder_cell.text().strip())
            self.inventory_service.update_item_sell_and_reorder(
                item_id=item_id,
                selling_price=selling_price,
                reorder_level=reorder_level,
                admin_pin=pin,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Inventory Edit", str(exc))
            self.refresh_inventory()
            return
        except Exception as exc:
            QMessageBox.critical(self, "Inventory Edit Error", str(exc))
            self.refresh_inventory()
            return

        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_reports()
        self.inventory_inline_status_label.setText(
            f"Saved inline edit for item #{item_id}: Sell={selling_price:.2f}, Reorder={reorder_level:.2f}"
        )
        self.inventory_inline_status_label.setVisible(True)
        QTimer.singleShot(2200, lambda: self.inventory_inline_status_label.setVisible(False))
        self._log_audit("inventory_inline_edit", "item", str(item_id), f"sell={selling_price}, reorder={reorder_level}")

    def refresh_purchases_tab(self) -> None:
        items = self.inventory_service.list_items()
        self.purchase_item_cache = {int(item["id"]): item for item in items}

        self.purchase_item_combo.clear()
        for item in items:
            self.purchase_item_combo.addItem(
                f"{item['name']} (Stock {item['stock_quantity']:.2f})",
                int(item["id"]),
            )
        self._on_purchase_item_changed()

        start_date, end_date = self._iso_range_from_edits(self.purchase_from_date, self.purchase_to_date)
        history = self.bookkeeping_service.list_purchases_between(
            start_date=start_date,
            end_date=end_date,
            limit=1000,
        )
        self.purchase_history_table.setRowCount(len(history))
        for row_index, purchase in enumerate(history):
            self.purchase_history_table.setItem(row_index, 0, QTableWidgetItem(str(purchase["id"])))
            self.purchase_history_table.setItem(
                row_index,
                1,
                QTableWidgetItem(purchase["supplier_name"] or "-"),
            )
            self.purchase_history_table.setItem(row_index, 2, QTableWidgetItem(purchase["purchased_at"]))
            lines_item = QTableWidgetItem(str(purchase["line_items"]))
            lines_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.purchase_history_table.setItem(row_index, 3, lines_item)
            self.purchase_history_table.setItem(
                row_index,
                4,
                QTableWidgetItem(f"{float(purchase['total_cost']):.2f}"),
            )
            total_item = self.purchase_history_table.item(row_index, 4)
            if total_item is not None:
                total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.purchase_history_table.setItem(
                row_index,
                5,
                QTableWidgetItem(purchase["notes"] or ""),
            )

        self.refresh_purchase_lines_table()

    def refresh_expenses_tab(self) -> None:
        self._updating_expense_table = True
        start_date, end_date = self._iso_range_from_edits(self.expense_from_date, self.expense_to_date)
        history = self.bookkeeping_service.list_expenses_between(
            start_date=start_date,
            end_date=end_date,
            limit=1000,
        )
        self.expense_history_table.setRowCount(len(history))
        for row_index, expense in enumerate(history):
            id_item = QTableWidgetItem(str(expense["id"]))
            id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)
            self.expense_history_table.setItem(row_index, 0, id_item)

            type_item = QTableWidgetItem(expense["expense_type"])
            self.expense_history_table.setItem(row_index, 1, type_item)

            amount_item = QTableWidgetItem(f"{float(expense['amount']):.2f}")
            self.expense_history_table.setItem(
                row_index,
                2,
                amount_item,
            )

            date_item = QTableWidgetItem(expense["spent_at"])
            date_item.setFlags(date_item.flags() & ~Qt.ItemIsEditable)
            self.expense_history_table.setItem(row_index, 3, date_item)

            notes_item = QTableWidgetItem(expense["notes"] or "")
            self.expense_history_table.setItem(row_index, 4, notes_item)
        self._updating_expense_table = False

    def _on_purchase_item_changed(self) -> None:
        item_id = self.purchase_item_combo.currentData()
        if item_id is None:
            return
        item = self.purchase_item_cache.get(int(item_id))
        if item is None:
            return
        self.purchase_cost_spin.setValue(float(item.get("cost_price") or 0))
        self._update_purchase_stock_preview()

    def _update_purchase_stock_preview(self) -> None:
        if not hasattr(self, "purchase_stock_preview_label"):
            return

        item_id = self.purchase_item_combo.currentData()
        if item_id is None:
            self.purchase_stock_preview_label.setText("")
            return

        item = self.purchase_item_cache.get(int(item_id))
        if item is None:
            self.purchase_stock_preview_label.setText("")
            return

        current_stock = float(item.get("stock_quantity") or 0)
        added_qty = float(self.purchase_qty_spin.value())
        projected_stock = current_stock + added_qty
        self.purchase_stock_preview_label.setText(
            f"Stock after purchase: {current_stock:.2f} -> {projected_stock:.2f}"
        )

    def add_purchase_line(self) -> None:
        item_id = self.purchase_item_combo.currentData()
        if item_id is None:
            QMessageBox.warning(self, "Purchase", "Please select an item.")
            return

        item = self.purchase_item_cache.get(int(item_id))
        if item is None:
            QMessageBox.warning(self, "Purchase", "Selected item was not found.")
            return

        quantity = float(self.purchase_qty_spin.value())
        cost_price = float(self.purchase_cost_spin.value())
        if quantity <= 0:
            QMessageBox.warning(self, "Purchase", "Quantity must be greater than zero.")
            return
        if cost_price <= 0:
            QMessageBox.warning(self, "Purchase", "Cost price must be greater than zero.")
            return

        existing = next((line for line in self.purchase_cart if int(line["item_id"]) == int(item_id)), None)
        if existing:
            existing["quantity"] = float(existing["quantity"]) + quantity
            existing["cost_price"] = cost_price
        else:
            self.purchase_cart.append(
                {
                    "item_id": int(item_id),
                    "name": item["name"],
                    "quantity": quantity,
                    "cost_price": cost_price,
                }
            )
        self.refresh_purchase_lines_table()
        self.purchase_qty_spin.setValue(1)
        self.purchase_item_combo.setFocus()
        if hasattr(self, "purchase_feedback_label"):
            self.purchase_feedback_label.setText(f"Added {item['name']} (Qty {quantity:.2f})")
            self.purchase_feedback_label.setVisible(True)
            QTimer.singleShot(1800, lambda: self.purchase_feedback_label.setVisible(False))

    def remove_selected_purchase_line(self) -> None:
        selected = self.purchase_lines_table.currentRow()
        if selected < 0:
            QMessageBox.information(self, "Purchase", "Please select a purchase line.")
            return

        name_item = self.purchase_lines_table.item(selected, 0)
        if name_item is None:
            return
        item_id = int(name_item.data(Qt.UserRole))
        self.purchase_cart = [line for line in self.purchase_cart if int(line["item_id"]) != item_id]
        self.refresh_purchase_lines_table()

    def clear_purchase_lines(self) -> None:
        self.purchase_cart = []
        self.refresh_purchase_lines_table()

    def _set_purchase_mode(self, editing: bool, purchase_id: int | None = None) -> None:
        if editing and purchase_id is not None:
            self.editing_purchase_id = int(purchase_id)
            self.save_purchase_btn.setText("Update Purchase")
            self.cancel_purchase_edit_btn.setVisible(True)
            self.cancel_purchase_edit_btn.setEnabled(True)
            self.purchase_mode_label.setText(f"Mode: Editing Purchase #{purchase_id}")
        else:
            self.editing_purchase_id = None
            self.save_purchase_btn.setText("Save Purchase")
            self.cancel_purchase_edit_btn.setVisible(False)
            self.cancel_purchase_edit_btn.setEnabled(False)
            self.purchase_mode_label.setText("Mode: New Purchase")

    def cancel_purchase_edit(self) -> None:
        self.purchase_supplier_input.clear()
        self.purchase_notes_input.clear()
        self.clear_purchase_lines()
        self._set_purchase_mode(False)

    def load_selected_purchase_for_edit(self) -> None:
        selected = self.purchase_history_table.currentRow()
        if selected < 0:
            QMessageBox.information(self, "Purchase Edit", "Please select a purchase from history.")
            return

        id_cell = self.purchase_history_table.item(selected, 0)
        if id_cell is None:
            return

        purchase_id = int(id_cell.text())
        pin = self._require_admin_access("Purchase Edit")
        if pin is None:
            return

        try:
            purchase = self.bookkeeping_service.get_purchase_for_edit(purchase_id=purchase_id, admin_pin=pin)
        except ValueError as exc:
            QMessageBox.warning(self, "Purchase Edit", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Purchase Edit Error", str(exc))
            return

        self.purchase_supplier_input.setText(purchase.get("supplier_name") or "")
        self.purchase_notes_input.setText(purchase.get("notes") or "")
        self.purchase_cart = [
            {
                "item_id": int(line["item_id"]),
                "name": line["name"],
                "quantity": float(line["quantity"]),
                "cost_price": float(line["cost_price"]),
            }
            for line in purchase.get("items", [])
        ]
        self.refresh_purchase_lines_table()
        self._set_purchase_mode(True, purchase_id=purchase_id)

    def refresh_purchase_lines_table(self) -> None:
        self._updating_purchase_table = True
        self.purchase_lines_table.setRowCount(len(self.purchase_cart))
        self.purchase_lines_table.setSortingEnabled(False)
        total = 0.0
        for row_index, line in enumerate(self.purchase_cart):
            line_total = float(line["quantity"]) * float(line["cost_price"])
            total += line_total
            name_item = QTableWidgetItem(line["name"])
            name_item.setData(Qt.UserRole, int(line["item_id"]))
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.purchase_lines_table.setItem(row_index, 0, name_item)

            qty_item = QTableWidgetItem(f"{line['quantity']:.2f}")
            qty_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.purchase_lines_table.setItem(row_index, 1, qty_item)

            cost_item = QTableWidgetItem(f"{line['cost_price']:.2f}")
            cost_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.purchase_lines_table.setItem(row_index, 2, cost_item)

            total_item = QTableWidgetItem(f"{line_total:.2f}")
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)
            self.purchase_lines_table.setItem(row_index, 3, total_item)

        self.purchase_lines_table.setSortingEnabled(True)
        self._updating_purchase_table = False
        self.purchase_empty_state_label.setVisible(len(self.purchase_cart) == 0)

        self.purchase_total_label.setText(f"TOTAL: INR {total:.2f}")

    def _on_purchase_line_item_changed(self, item: QTableWidgetItem) -> None:
        if self._updating_purchase_table:
            return

        row = item.row()
        column = item.column()
        if column not in (1, 2):
            return

        name_cell = self.purchase_lines_table.item(row, 0)
        if name_cell is None:
            return

        item_id = int(name_cell.data(Qt.UserRole))
        target = next((line for line in self.purchase_cart if int(line["item_id"]) == item_id), None)
        if target is None:
            return

        try:
            value = float(item.text().strip())
            if column == 1 and value <= 0:
                raise ValueError("Quantity must be greater than zero.")
            if column == 2 and value <= 0:
                raise ValueError("Cost must be greater than zero.")
        except ValueError as exc:
            QMessageBox.warning(self, "Purchase Line", str(exc))
            self.refresh_purchase_lines_table()
            return

        if column == 1:
            target["quantity"] = value
        else:
            target["cost_price"] = value
        self.refresh_purchase_lines_table()

    def save_purchase(self) -> None:
        if not self.purchase_cart:
            QMessageBox.warning(self, "Purchase", "Add at least one purchase line.")
            return

        was_editing = self.editing_purchase_id is not None

        supplier_name = self.purchase_supplier_input.text().strip()
        notes = self.purchase_notes_input.text().strip()
        payload = [
            {
                "item_id": int(line["item_id"]),
                "quantity": float(line["quantity"]),
                "cost_price": float(line["cost_price"]),
            }
            for line in self.purchase_cart
        ]

        try:
            if self.editing_purchase_id is None:
                purchase_id = self.bookkeeping_service.add_purchase(
                    supplier_name=supplier_name,
                    items=payload,
                    notes=notes,
                )
                message = f"Purchase recorded. ID: {purchase_id}"
            else:
                pin = self._require_admin_access("Purchase Update")
                if pin is None:
                    return
                self.bookkeeping_service.update_purchase(
                    purchase_id=self.editing_purchase_id,
                    supplier_name=supplier_name,
                    items=payload,
                    notes=notes,
                    admin_pin=pin,
                )
                purchase_id = self.editing_purchase_id
                message = f"Purchase updated. ID: {purchase_id}"
        except ValueError as exc:
            QMessageBox.warning(self, "Purchase", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Purchase Error", str(exc))
            return

        QMessageBox.information(self, "Purchase Saved", message)
        self.purchase_supplier_input.clear()
        self.purchase_notes_input.clear()
        self.clear_purchase_lines()
        self._set_purchase_mode(False)
        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_purchases_tab()
        self.refresh_reports()
        self._log_audit(
            "purchase_update" if was_editing else "purchase_save",
            "purchase",
            str(purchase_id),
            f"lines={len(payload)}",
        )

    def add_expense(self) -> None:
        expense_type = self.expense_type_combo.currentText().strip()
        amount = float(self.expense_amount_spin.value())
        notes = self.expense_notes_input.text().strip()

        try:
            expense_id = self.bookkeeping_service.add_expense(
                expense_type=expense_type,
                amount=amount,
                notes=notes,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Expense", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Expense Error", str(exc))
            return

        QMessageBox.information(self, "Expense Saved", f"Expense recorded. ID: {expense_id}")
        self.expense_amount_spin.setValue(0.01)
        self.expense_notes_input.clear()
        self.refresh_expenses_tab()
        self.refresh_reports()
        self._log_audit("expense_add", "expense", str(expense_id), expense_type)

    def _on_expense_item_changed(self, item: QTableWidgetItem) -> None:
        if self._updating_expense_table:
            return

        row = item.row()
        column = item.column()
        if column not in (1, 2, 4):
            return

        id_cell = self.expense_history_table.item(row, 0)
        type_cell = self.expense_history_table.item(row, 1)
        amount_cell = self.expense_history_table.item(row, 2)
        notes_cell = self.expense_history_table.item(row, 4)
        if id_cell is None or type_cell is None or amount_cell is None or notes_cell is None:
            return

        try:
            expense_id = int(id_cell.text())
            expense_type = type_cell.text().strip()
            amount = float(amount_cell.text().strip())
            notes = notes_cell.text().strip()
            self.bookkeeping_service.update_expense(
                expense_id=expense_id,
                expense_type=expense_type,
                amount=amount,
                notes=notes,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Expense Edit", str(exc))
            self.refresh_expenses_tab()
            return
        except Exception as exc:
            QMessageBox.critical(self, "Expense Edit Error", str(exc))
            self.refresh_expenses_tab()
            return

        self.refresh_expenses_tab()
        self.refresh_reports()
        self._log_audit("expense_edit", "expense", str(expense_id), f"type={expense_type}, amount={amount:.2f}")

    def refresh_billing_items(self) -> None:
        items = self.inventory_service.list_items()
        self.billing_items_cache = [i for i in items if (i.get("item_kind") or "sellable") == "sellable"]
        self.apply_billing_filter()
        self.refresh_cigarette_quick_buttons()
        self._update_billing_dashboard_metrics()

    def refresh_cigarette_quick_buttons(self) -> None:
        grouped = self.inventory_service.cigarette_items_grouped()

        self._clear_layout(self.small_buttons_layout)
        self._clear_layout(self.medium_buttons_layout)
        self._clear_layout(self.big_buttons_layout)
        for shortcut in self.cigarette_shortcuts:
            shortcut.setParent(None)
        self.cigarette_shortcuts = []

        shortcut_number = 1
        shortcut_number = self._populate_quick_row(
            layout=self.small_buttons_layout,
            items=grouped.get("small", []),
            color="#2e8f56",
            start_shortcut=shortcut_number,
        )
        shortcut_number = self._populate_quick_row(
            layout=self.medium_buttons_layout,
            items=grouped.get("medium", []),
            color="#cf8a21",
            start_shortcut=shortcut_number,
        )
        self._populate_quick_row(
            layout=self.big_buttons_layout,
            items=grouped.get("big", []),
            color="#b44b4b",
            start_shortcut=shortcut_number,
        )

    def _populate_quick_row(
        self,
        layout: QHBoxLayout,
        items: list[dict],
        color: str,
        start_shortcut: int,
    ) -> int:
        shortcut_number = start_shortcut
        for item in items:
            label = f"{item['name']} {item['selling_price']:.0f}"
            if 1 <= shortcut_number <= 9:
                label += f" ({shortcut_number})"
            btn = QPushButton(label)
            btn.setStyleSheet(
                f"background-color: {color};"
                "color: white;"
                "font-weight: 700;"
                "font-size: 13px;"
                "border-radius: 10px;"
                "padding: 8px 11px;"
                "min-height: 34px;"
                "border: 1px solid rgba(255,255,255,0.18);"
            )
            btn.clicked.connect(lambda _, item_id=item["id"]: self.add_item_to_cart_by_id(item_id))
            layout.addWidget(btn)

            if 1 <= shortcut_number <= 9:
                sc = QShortcut(QKeySequence(str(shortcut_number)), self)
                sc.activated.connect(lambda item_id=item["id"]: self.add_item_to_cart_by_id(item_id))
                self.cigarette_shortcuts.append(sc)
            shortcut_number += 1

        layout.addStretch()
        return shortcut_number

    @staticmethod
    def _clear_layout(layout: QHBoxLayout) -> None:
        while layout.count():
            child = layout.takeAt(0)
            widget = child.widget()
            if widget is not None:
                widget.deleteLater()

    def apply_billing_filter(self) -> None:
        query = self.search_input.text().strip().lower()
        if query:
            filtered = [i for i in self.billing_items_cache if query in i["name"].lower()]
        else:
            filtered = self.billing_items_cache

        self._filtered_billing_count = len(filtered)

        table = self.billing_items_table
        table.setRowCount(len(filtered))

        for row_index, item in enumerate(filtered):
            name_item = QTableWidgetItem(item["name"])
            name_item.setData(Qt.UserRole, int(item["id"]))
            table.setItem(row_index, 0, name_item)
            table.setItem(row_index, 1, QTableWidgetItem(f"{item['selling_price']:.2f}"))
            table.setItem(row_index, 2, QTableWidgetItem(f"{item['stock_quantity']:.2f}"))

        if len(filtered) > 0:
            table.selectRow(0)
            self.billing_empty_state_label.setVisible(False)
        else:
            if query:
                self.billing_empty_state_label.setText(
                    f"No items match '{query}'. Try a broader keyword or click Refresh Items."
                )
            else:
                self.billing_empty_state_label.setText(
                    "No billing items available. Add stock in Inventory or click Refresh Items."
                )
            self.billing_empty_state_label.setVisible(True)

        self._update_billing_dashboard_metrics()

    def _update_billing_dashboard_metrics(self, total_amount: float | None = None) -> None:
        if not hasattr(self, "billing_cart_lines_value"):
            return

        cart_lines = len(self.cart)
        cart_qty = sum(float(item.get("quantity", 0)) for item in self.cart.values())

        self.billing_cart_lines_value.setText(str(cart_lines))
        self.billing_cart_qty_value.setText(f"{cart_qty:.2f}")

        if total_amount is None:
            total_text = self.total_label.text().replace("TOTAL: INR ", "").strip()
            try:
                total_amount = float(total_text)
            except ValueError:
                total_amount = 0.0
        self.billing_total_value.setText(f"INR {total_amount:.2f}")
        self.total_label.setText(f"TOTAL: INR {total_amount:.2f}")

    def _on_item_kind_changed(self) -> None:
        kind = self.item_kind_combo.currentData() if hasattr(self, "item_kind_combo") else "sellable"
        if kind == "ingredient":
            self.costing_mode_combo.setCurrentIndex(self.costing_mode_combo.findData("manual"))
            self.costing_mode_combo.setEnabled(False)
            self.stock_tracked_checkbox.setChecked(True)
            self.stock_tracked_checkbox.setEnabled(False)
        else:
            self.costing_mode_combo.setEnabled(True)
            self.stock_tracked_checkbox.setEnabled(True)

    def add_inventory_item(self) -> None:
        try:
            self.inventory_service.add_item(
                name=self.item_name_input.text(),
                category_id=self.category_combo.currentData(),
                selling_price=float(self.sell_price_spin.value()),
                cost_price=float(self.cost_price_spin.value()),
                stock_quantity=float(self.stock_spin.value()),
                reorder_level=float(self.reorder_spin.value()),
                item_kind=str(self.item_kind_combo.currentData() or "sellable"),
                costing_mode=str(self.costing_mode_combo.currentData() or "manual"),
                unit_name=self.unit_name_input.text().strip() or "pcs",
                is_stock_tracked=self.stock_tracked_checkbox.isChecked(),
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Validation Error", str(exc))
            return

        self.item_name_input.clear()
        self.sell_price_spin.setValue(0)
        self.cost_price_spin.setValue(0)
        self.stock_spin.setValue(0)
        self.reorder_spin.setValue(0)
        self.item_kind_combo.setCurrentIndex(self.item_kind_combo.findData("sellable"))
        self.costing_mode_combo.setCurrentIndex(self.costing_mode_combo.findData("manual"))
        self.unit_name_input.setText("pcs")
        self.stock_tracked_checkbox.setChecked(True)
        self._on_item_kind_changed()
        self._sync_inventory_add_button_state()
        self._focus_inventory_name()
        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_reports()
        if hasattr(self, "inventory_inline_status_label"):
            self.inventory_inline_status_label.setText("Item added successfully.")
            self.inventory_inline_status_label.setVisible(True)
            QTimer.singleShot(2200, lambda: self.inventory_inline_status_label.setVisible(False))

    def _selected_inventory_item(self) -> dict | None:
        selected = self.inventory_items_table.currentRow()
        if selected < 0:
            QMessageBox.information(self, "Select Item", "Please select an item from inventory.")
            return None

        item_id = self._inventory_id_for_row(selected)
        if item_id is None:
            QMessageBox.warning(self, "Select Item", "Could not resolve selected inventory item.")
            return None

        return {
            "item_id": item_id,
            "name": self.inventory_items_table.item(selected, 0).text(),
            "selling_price": float(self.inventory_items_table.item(selected, 4).text()),
        }

    def _log_audit(
        self,
        action_type: str,
        entity_type: str = "",
        entity_id: str = "",
        details: str = "",
    ) -> None:
        try:
            self.bookkeeping_service.log_audit(
                actor_role=self.current_role,
                action_type=action_type,
                entity_type=entity_type,
                entity_id=entity_id,
                details=details,
            )
        except Exception:
            pass

    def _ask_admin_pin(self) -> str | None:
        pin, ok = QInputDialog.getText(self, "Admin PIN", "Enter admin PIN:")
        if not ok:
            return None
        return pin.strip()

    def _require_admin_access(self, action_name: str) -> str | None:
        if self.current_role == "admin":
            return self.bookkeeping_service.get_setting("admin_pin", "1234") or "1234"

        pin = self._ask_admin_pin()
        if pin is None:
            return None
        if not self.bookkeeping_service.verify_admin_pin(pin):
            QMessageBox.warning(self, action_name, "Invalid admin PIN.")
            return None
        return pin

    def on_role_changed(self, role_name: str) -> None:
        role = role_name.strip().lower()
        if role == self.current_role:
            return

        if role == "admin":
            pin = self._ask_admin_pin()
            if pin is None or not self.bookkeeping_service.verify_admin_pin(pin):
                QMessageBox.warning(self, "Role Switch", "Admin PIN verification failed.")
                self.role_combo.blockSignals(True)
                self.role_combo.setCurrentText(self.current_role)
                self.role_combo.blockSignals(False)
                return

        self.current_role = role
        self.bookkeeping_service.set_setting("current_role", role)
        self._log_audit("role_switch", "session", "current", f"Role changed to {role}")

    def _configure_auto_backup_timer(self) -> None:
        enabled = (self.bookkeeping_service.get_setting("auto_backup_enabled", "0") or "0") == "1"
        interval_str = self.bookkeeping_service.get_setting("backup_interval_minutes", "60") or "60"
        try:
            interval_minutes = max(5, int(float(interval_str)))
        except ValueError:
            interval_minutes = 60

        if hasattr(self, "auto_backup_enabled_checkbox"):
            self.auto_backup_enabled_checkbox.setChecked(enabled)
        if hasattr(self, "auto_backup_interval_spin"):
            self.auto_backup_interval_spin.setValue(interval_minutes)

        self.auto_backup_timer.stop()
        if enabled:
            self.auto_backup_timer.setInterval(interval_minutes * 60 * 1000)
            self.auto_backup_timer.start()

    def save_backup_preferences(self) -> None:
        enabled = self.auto_backup_enabled_checkbox.isChecked()
        interval = int(self.auto_backup_interval_spin.value())
        self.bookkeeping_service.set_setting("auto_backup_enabled", "1" if enabled else "0")
        self.bookkeeping_service.set_setting("backup_interval_minutes", str(interval))
        self._configure_auto_backup_timer()
        self._log_audit(
            "backup_schedule_update",
            "settings",
            "auto_backup",
            f"enabled={enabled}, interval_minutes={interval}",
        )
        QMessageBox.information(self, "Backup Schedule", "Backup schedule updated.")

    def _run_scheduled_backup(self) -> None:
        try:
            backup_path = create_backup(db_path=self.db_path)
            self._log_audit("scheduled_backup", "backup", str(backup_path), "Automatic backup completed")
        except Exception as exc:
            self._log_audit("scheduled_backup_failed", "backup", "", str(exc))

    @staticmethod
    def _format_restore_preview(current_counts: dict, backup_counts: dict, backup_file: str) -> str:
        lines = [
            "Restoring will overwrite current live database.",
            "",
            f"Backup file: {backup_file}",
            "",
            "Current DB -> Backup DB",
            f"Items: {current_counts.get('items', 0)} -> {backup_counts.get('items', 0)}",
            f"Sales: {current_counts.get('sales', 0)} -> {backup_counts.get('sales', 0)}",
            f"Purchases: {current_counts.get('purchases', 0)} -> {backup_counts.get('purchases', 0)}",
            f"Expenses: {current_counts.get('expenses', 0)} -> {backup_counts.get('expenses', 0)}",
            f"Stock Movements: {current_counts.get('stock_movements', 0)} -> {backup_counts.get('stock_movements', 0)}",
            "",
            "Continue with restore?",
        ]
        return "\n".join(lines)

    def export_reports_xlsx(self) -> None:
        try:
            from openpyxl import Workbook
        except Exception:
            QMessageBox.warning(self, "XLSX Export", "openpyxl is not installed. Run: pip install openpyxl")
            return

        start_date, end_date = self._iso_range_from_edits(self.report_from_date, self.report_to_date)
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save XLSX",
            f"reports_{self._range_suffix(start_date, end_date)}.xlsx",
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["From", start_date])
        ws_summary.append(["To", end_date])
        ws_summary.append(["Sales", self.sales_value.text().replace("INR ", "")])
        ws_summary.append(["COGS", self.cogs_value.text().replace("INR ", "")])
        ws_summary.append(["Gross Profit", self.gross_profit_value.text().replace("INR ", "")])
        ws_summary.append(["Purchases", self.purchases_value.text().replace("INR ", "")])
        ws_summary.append(["Expenses", self.expenses_value.text().replace("INR ", "")])
        ws_summary.append(["Fixed Cost / Day", self.fixed_daily_value.text().replace("INR ", "")])
        ws_summary.append(["Realistic Daily Profit", self.net_value.text().replace("INR ", "")])

        table_sheets = [
            ("SalesTrend", self.sales_trend_table),
            ("TopItems", self.top_items_table),
            ("LowStock", self.low_stock_table),
            ("StockLedger", self.ledger_table),
        ]
        for sheet_name, table in table_sheets:
            ws = wb.create_sheet(title=sheet_name)
            for row in self._table_rows_for_csv(table):
                ws.append(row)

        try:
            wb.save(path)
        except Exception as exc:
            QMessageBox.critical(self, "XLSX Export", f"Failed to save XLSX: {exc}")
            return

        self._log_audit("xlsx_export", "report", "range", f"{start_date}..{end_date}")
        if self.open_after_export_checkbox.isChecked() and hasattr(os, "startfile"):
            try:
                os.startfile(path)
            except Exception:
                pass
        QMessageBox.information(self, "XLSX Export", f"XLSX exported successfully:\n{path}")

    def export_printable_summary(self) -> None:
        start_date, end_date = self._iso_range_from_edits(self.report_from_date, self.report_to_date)
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Printable Summary",
            f"summary_{self._range_suffix(start_date, end_date)}.txt",
            "Text Files (*.txt)",
        )
        if not path:
            return

        lines = [
            "Cafe POS - Printable Summary",
            f"Range: {start_date} to {end_date}",
            "",
            f"Sales: {self.sales_value.text()}",
            f"COGS: {self.cogs_value.text()}",
            f"Gross Profit: {self.gross_profit_value.text()}",
            f"Purchases: {self.purchases_value.text()}",
            f"Expenses: {self.expenses_value.text()}",
            f"Fixed Cost / Day: {self.fixed_daily_value.text()}",
            f"Realistic Daily Profit: {self.net_value.text()}",
            "",
            "Top Selling Items:",
        ]
        for row in range(min(self.top_items_table.rowCount(), 15)):
            item = self.top_items_table.item(row, 0).text() if self.top_items_table.item(row, 0) else ""
            qty = self.top_items_table.item(row, 1).text() if self.top_items_table.item(row, 1) else ""
            value = self.top_items_table.item(row, 2).text() if self.top_items_table.item(row, 2) else ""
            lines.append(f"- {item}: qty {qty}, value {value}")

        try:
            Path(path).write_text("\n".join(lines), encoding="utf-8")
        except Exception as exc:
            QMessageBox.critical(self, "Printable Summary", f"Failed: {exc}")
            return

        self._log_audit("printable_summary_export", "report", "range", f"{start_date}..{end_date}")
        if self.open_after_export_checkbox.isChecked() and hasattr(os, "startfile"):
            try:
                os.startfile(path)
            except Exception:
                pass
        QMessageBox.information(self, "Printable Summary", f"Summary exported:\n{path}")

    def export_audit_csv(self) -> None:
        if not hasattr(self, "audit_log_table"):
            return
        rows = self._table_rows_for_csv(self.audit_log_table)
        self._save_csv_rows("audit_log_export.csv", rows)
        self._log_audit("audit_export_csv", "audit_log", "table", "audit log exported as csv")

    def export_audit_xlsx(self) -> None:
        try:
            from openpyxl import Workbook
        except Exception:
            QMessageBox.warning(self, "Audit XLSX", "openpyxl is not installed. Run: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Audit XLSX",
            "audit_log_export.xlsx",
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "AuditLog"
        for row in self._table_rows_for_csv(self.audit_log_table):
            ws.append(row)

        try:
            wb.save(path)
        except Exception as exc:
            QMessageBox.critical(self, "Audit XLSX", f"Failed to save XLSX: {exc}")
            return

        self._log_audit("audit_export_xlsx", "audit_log", "table", "audit log exported as xlsx")
        if self.open_after_export_checkbox.isChecked() and hasattr(os, "startfile"):
            try:
                os.startfile(path)
            except Exception:
                pass
        QMessageBox.information(self, "Audit XLSX", f"Audit XLSX exported successfully:\n{path}")

    def manage_selected_item_recipe(self) -> None:
        item = self._selected_inventory_item()
        if item is None:
            return

        selected_full = next(
            (row for row in self.inventory_items_cache if int(row["id"]) == int(item["item_id"])),
            None,
        )
        if selected_full is None:
            QMessageBox.warning(self, "Recipe", "Selected item could not be resolved.")
            return
        if (selected_full.get("item_kind") or "sellable") != "sellable":
            QMessageBox.warning(self, "Recipe", "Recipes can be configured only for sellable items.")
            return

        pin = self._require_admin_access("Manage Recipe")
        if pin is None:
            return

        ingredients = self.inventory_service.list_ingredients()
        if not ingredients:
            QMessageBox.warning(self, "Recipe", "No ingredient items found. Create ingredients first.")
            return

        existing = self.inventory_service.get_recipe(int(item["item_id"]))

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Recipe - {item['name']}")
        dialog.resize(760, 460)
        layout = QVBoxLayout(dialog)

        header = QLabel("Define ingredient usage per recipe batch. Sale-time cost will be derived from these lines.")
        header.setWordWrap(True)
        layout.addWidget(header)

        meta_form = QFormLayout()
        yield_spin = QDoubleSpinBox()
        yield_spin.setDecimals(3)
        yield_spin.setMinimum(0.001)
        yield_spin.setMaximum(100000)
        yield_spin.setValue(float(existing.get("yield_qty", 1.0)) if existing else 1.0)
        meta_form.addRow("Recipe Yield Qty", yield_spin)
        layout.addLayout(meta_form)

        table = QTableWidget(0, 3)
        table.setHorizontalHeaderLabels(["Ingredient", "Qty Used", "Waste %"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        layout.addWidget(table)

        def add_line(
            ingredient_item_id: int | None = None,
            quantity_used: float = 1.0,
            waste_percent: float = 0.0,
        ) -> None:
            row = table.rowCount()
            table.insertRow(row)

            ingredient_combo = QComboBox()
            for ing in ingredients:
                ingredient_combo.addItem(
                    f"{ing['name']} ({ing.get('unit_name') or 'unit'})",
                    int(ing["id"]),
                )
            if ingredient_item_id is not None:
                idx = ingredient_combo.findData(int(ingredient_item_id))
                if idx >= 0:
                    ingredient_combo.setCurrentIndex(idx)
            table.setCellWidget(row, 0, ingredient_combo)

            qty_spin = QDoubleSpinBox()
            qty_spin.setDecimals(4)
            qty_spin.setMinimum(0.0001)
            qty_spin.setMaximum(100000)
            qty_spin.setValue(float(quantity_used))
            table.setCellWidget(row, 1, qty_spin)

            waste_spin = QDoubleSpinBox()
            waste_spin.setDecimals(2)
            waste_spin.setMinimum(0)
            waste_spin.setMaximum(100)
            waste_spin.setValue(float(waste_percent))
            table.setCellWidget(row, 2, waste_spin)

        if existing and existing.get("lines"):
            for line in existing["lines"]:
                add_line(
                    ingredient_item_id=int(line["ingredient_item_id"]),
                    quantity_used=float(line["quantity_used"]),
                    waste_percent=float(line.get("waste_percent", 0.0)),
                )
        else:
            add_line()

        line_actions = QHBoxLayout()
        add_btn = QPushButton("Add Ingredient Line")
        remove_btn = QPushButton("Remove Selected Line")
        add_btn.clicked.connect(lambda: add_line())
        remove_btn.clicked.connect(
            lambda: table.removeRow(table.currentRow()) if table.currentRow() >= 0 else None
        )
        line_actions.addWidget(add_btn)
        line_actions.addWidget(remove_btn)
        line_actions.addStretch()
        layout.addLayout(line_actions)

        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() != QDialog.Accepted:
            return

        lines: list[dict] = []
        for row in range(table.rowCount()):
            ingredient_combo = table.cellWidget(row, 0)
            qty_spin = table.cellWidget(row, 1)
            waste_spin = table.cellWidget(row, 2)
            if not isinstance(ingredient_combo, QComboBox):
                continue
            if not isinstance(qty_spin, QDoubleSpinBox):
                continue
            if not isinstance(waste_spin, QDoubleSpinBox):
                continue
            lines.append(
                {
                    "ingredient_item_id": int(ingredient_combo.currentData()),
                    "quantity_used": float(qty_spin.value()),
                    "waste_percent": float(waste_spin.value()),
                }
            )

        if not lines:
            QMessageBox.warning(self, "Recipe", "Recipe must include at least one ingredient line.")
            return

        try:
            self.inventory_service.save_recipe(
                sellable_item_id=int(item["item_id"]),
                lines=lines,
                yield_qty=float(yield_spin.value()),
                admin_pin=pin,
            )
            self.inventory_service.set_item_classification(
                item_id=int(item["item_id"]),
                item_kind="sellable",
                costing_mode="recipe",
                is_stock_tracked=bool(int(selected_full.get("is_stock_tracked", 1))),
                unit_name=(selected_full.get("unit_name") or "pcs"),
                admin_pin=pin,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Recipe", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Recipe Error", str(exc))
            return

        QMessageBox.information(self, "Recipe", "Recipe saved successfully.")
        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_reports()
        self._log_audit("recipe_save", "item", str(item["item_id"]), f"lines={len(lines)}")

    def update_selected_item_price(self) -> None:
        item = self._selected_inventory_item()
        if item is None:
            return

        pin = self._require_admin_access("Update Price")
        if pin is None:
            return

        new_sell, ok_sell = QInputDialog.getDouble(
            self,
            "Update Selling Price",
            f"New selling price for {item['name']}",
            value=item["selling_price"],
            minValue=0.01,
            decimals=2,
        )
        if not ok_sell:
            return

        new_cost, ok_cost = QInputDialog.getDouble(
            self,
            "Update Cost Price",
            f"New cost price for {item['name']}",
            value=0,
            minValue=0,
            decimals=2,
        )
        if not ok_cost:
            return

        try:
            self.inventory_service.update_item_pricing(
                item_id=item["item_id"],
                selling_price=float(new_sell),
                cost_price=float(new_cost),
                admin_pin=pin,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Update Failed", str(exc))
            return

        self.refresh_inventory()
        self.refresh_billing_items()
        self._log_audit("price_update", "item", str(item["item_id"]), f"sell={float(new_sell):.2f}, cost={float(new_cost):.2f}")

    def adjust_selected_item_stock(self) -> None:
        item = self._selected_inventory_item()
        if item is None:
            return

        pin = self._require_admin_access("Manual Stock Adjustment")
        if pin is None:
            return

        quantity_delta, ok_delta = QInputDialog.getDouble(
            self,
            "Manual Stock Adjustment",
            "Enter stock change (+ add, - reduce):",
            decimals=2,
        )
        if not ok_delta:
            return

        notes, ok_notes = QInputDialog.getText(self, "Reason", "Reason for adjustment:")
        if not ok_notes:
            return

        try:
            self.inventory_service.manual_stock_adjustment(
                item_id=item["item_id"],
                quantity_delta=float(quantity_delta),
                admin_pin=pin,
                notes=notes,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Adjustment Failed", str(exc))
            return

        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_reports()
        self._log_audit("manual_stock_adjust", "item", str(item["item_id"]), f"delta={float(quantity_delta):.2f}, notes={notes}")

    def load_starter_cigarettes(self) -> None:
        try:
            created = self.inventory_service.load_starter_cigarette_items()
        except ValueError as exc:
            QMessageBox.warning(self, "Starter Data", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Starter Data Error", str(exc))
            return

        QMessageBox.information(
            self,
            "Starter Data Loaded",
            f"Created {created} cigarette SKUs for quick billing presets.",
        )
        self.refresh_inventory()
        self.refresh_billing_items()

    def delete_selected_item(self) -> None:
        item = self._selected_inventory_item()
        if item is None:
            return

        confirm = QMessageBox.question(
            self,
            "Delete Item",
            f"Delete item '{item['name']}' from active inventory?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        pin = self._require_admin_access("Delete Item")
        if pin is None:
            return

        try:
            self.inventory_service.delete_item(item_id=item["item_id"], admin_pin=pin)
        except ValueError as exc:
            QMessageBox.warning(self, "Delete Failed", str(exc))
            return

        QMessageBox.information(self, "Item Deleted", f"{item['name']} was removed from active inventory.")
        self.refresh_inventory()
        self.refresh_billing_items()
        self.refresh_reports()
        self._log_audit("item_delete", "item", str(item["item_id"]), item["name"])

    def _on_catalog_double_click(self, row: int, _: int) -> None:
        self.billing_items_table.selectRow(row)

    def _on_catalog_single_click(self, row: int, _: int) -> None:
        self.billing_items_table.selectRow(row)
        self.add_selected_item_to_cart()

    def add_item_to_cart_by_id(self, item_id: int, quantity: float | None = None) -> None:
        item = next((i for i in self.billing_items_cache if int(i["id"]) == int(item_id)), None)
        if item is None:
            QMessageBox.warning(self, "Item Missing", "Selected item is no longer available.")
            return

        qty = float(quantity if quantity is not None else self.qty_spin.value())
        existing_qty = float(self.cart.get(item_id, {}).get("quantity", 0))
        stock_tracked = int(item.get("is_stock_tracked", 1)) == 1
        recipe_costed = (item.get("costing_mode") or "manual") == "recipe"
        if stock_tracked and not recipe_costed:
            available_stock = float(item["stock_quantity"])
            if qty + existing_qty > available_stock:
                QMessageBox.warning(
                    self,
                    "Insufficient Stock",
                    f"Requested quantity exceeds available stock ({available_stock:.2f}).",
                )
                return

        self.cart[item_id] = {
            "item_id": item_id,
            "name": item["name"],
            "quantity": qty + existing_qty,
            "unit_price": float(item["selling_price"]),
        }
        self.refresh_cart_table()
        QApplication.beep()

    def add_selected_item_to_cart(self) -> None:
        selected = self.billing_items_table.currentRow()
        if selected < 0:
            QMessageBox.information(self, "Select Item", "Please select an item first.")
            return

        item_cell = self.billing_items_table.item(selected, 0)
        if item_cell is None:
            return
        item_id = int(item_cell.data(Qt.UserRole))
        self.add_item_to_cart_by_id(item_id)

    def _selected_cart_item_id(self) -> int | None:
        selected = self.cart_table.currentRow()
        if selected < 0:
            QMessageBox.information(self, "Select Cart Item", "Please select an item in the cart.")
            return None

        return int(self.cart_table.item(selected, 0).text())

    def remove_selected_cart_item(self) -> None:
        item_id = self._selected_cart_item_id()
        if item_id is None:
            return

        self.cart.pop(item_id, None)
        self.refresh_cart_table()

    def increase_selected_cart_item_qty(self) -> None:
        item_id = self._selected_cart_item_id()
        if item_id is None:
            return

        self.add_item_to_cart_by_id(item_id, quantity=float(self.qty_spin.value()))

    def decrease_selected_cart_item_qty(self) -> None:
        item_id = self._selected_cart_item_id()
        if item_id is None:
            return

        line = self.cart.get(item_id)
        if line is None:
            return

        next_qty = float(line["quantity"]) - float(self.qty_spin.value())
        if next_qty <= 0:
            self.cart.pop(item_id, None)
        else:
            line["quantity"] = next_qty
            self.cart[item_id] = line

        self.refresh_cart_table()

    def refresh_cart_table(self) -> None:
        items = list(self.cart.values())
        self.cart_table.setRowCount(len(items))

        total = 0.0
        for row_index, item in enumerate(items):
            line_total = item["quantity"] * item["unit_price"]
            total += line_total

            self.cart_table.setItem(row_index, 0, QTableWidgetItem(str(item["item_id"])))
            self.cart_table.setItem(row_index, 1, QTableWidgetItem(item["name"]))
            self.cart_table.setItem(row_index, 2, QTableWidgetItem(f"{item['quantity']:.2f}"))
            self.cart_table.setItem(row_index, 3, QTableWidgetItem(f"{item['unit_price']:.2f}"))
            self.cart_table.setItem(row_index, 4, QTableWidgetItem(f"{line_total:.2f}"))

        self._update_billing_dashboard_metrics(total_amount=total)

    def clear_cart(self) -> None:
        self.cart.clear()
        self.refresh_cart_table()
        if self.cart_file.exists():
            self.cart_file.unlink(missing_ok=True)

    def checkout(self) -> None:
        cart_items = [
            {"item_id": item["item_id"], "quantity": item["quantity"]}
            for item in self.cart.values()
        ]

        try:
            sale_result = self.sales_service.checkout(cart_items)
            sale = self.sales_service.sale_details(int(sale_result["sale_id"]))
        except ValueError as exc:
            QMessageBox.warning(self, "Checkout Error", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Unexpected Error", str(exc))
            return

        sale["total_amount"] = sale_result["total_amount"]
        printed_path = self._print_with_retry(sale)

        message = (
            f"Bill saved successfully.\n"
            f"Invoice: {sale_result['invoice_number']}\n"
            f"Sale ID: {sale_result['sale_id']}"
        )
        if printed_path:
            message += f"\nPrinted copy: {printed_path}"
        else:
            message += "\nBill saved but print skipped/failed."

        QMessageBox.information(self, "Checkout Complete", message)
        self.clear_cart()
        self.refresh_billing_items()
        self.refresh_inventory()
        self.refresh_reports()
        self._log_audit("sale_checkout", "sale", str(sale_result["sale_id"]), sale_result["invoice_number"])

    def _print_with_retry(self, sale_payload: dict) -> str | None:
        while True:
            try:
                path = self.print_service.print_bill(sale_payload)
                return str(path)
            except Exception as exc:
                response = QMessageBox.question(
                    self,
                    "Printer Error",
                    f"Bill is saved, but printing failed: {exc}\nRetry printing?",
                    QMessageBox.Yes | QMessageBox.No,
                )
                if response == QMessageBox.No:
                    return None

    def refresh_reports(self) -> None:
        start_date, end_date = self._iso_range_from_edits(self.report_from_date, self.report_to_date)
        summary = self.report_service.summary_between(start_date=start_date, end_date=end_date)
        self.sales_value.setText(f"INR {summary['sales']:.2f}")
        self.cogs_value.setText(f"INR {summary['cogs']:.2f}")
        self.gross_profit_value.setText(f"INR {summary['gross_profit']:.2f}")
        self.purchases_value.setText(f"INR {summary['purchases']:.2f}")
        self.expenses_value.setText(f"INR {summary['expenses']:.2f}")
        self.fixed_daily_value.setText(f"INR {summary['daily_fixed_overhead']:.2f}")
        self.net_value.setText(f"INR {summary['net_profit']:.2f}")

        fixed = summary.get("fixed_costs", {})
        self.rent_spin.setValue(float(fixed.get("rent", 0)))
        self.salary_spin.setValue(float(fixed.get("salary", 0)))
        self.maintenance_spin.setValue(float(fixed.get("maintenance", 0)))
        self.electricity_spin.setValue(float(fixed.get("electricity", 0)))
        self.monthly_fixed_total_label.setText(
            f"Monthly Fixed Total: INR {summary['monthly_fixed_total']:.2f}"
        )

        if hasattr(self, "overhead_date_edit"):
            overhead_date = self.overhead_date_edit.date().toString("yyyy-MM-dd")
            overhead = self.report_service.daily_overhead(overhead_date)
            self.overhead_gas_spin.setValue(float(overhead.get("gas_cost", 0)))
            self.overhead_labor_spin.setValue(float(overhead.get("labor_cost", 0)))
            self.overhead_misc_spin.setValue(float(overhead.get("misc_cost", 0)))
            self.overhead_units_spin.setValue(float(overhead.get("expected_units", 0)))
            total_overhead = (
                float(overhead.get("gas_cost", 0))
                + float(overhead.get("labor_cost", 0))
                + float(overhead.get("misc_cost", 0))
            )
            expected_units = float(overhead.get("expected_units", 0))
            per_unit = total_overhead / expected_units if expected_units > 0 else 0.0
            self.overhead_per_unit_label.setText(f"Overhead / Unit: INR {per_unit:.2f}")

        low_stock = self.report_service.low_stock()
        self.low_stock_table.setRowCount(len(low_stock))
        for row_index, item in enumerate(low_stock):
            self.low_stock_table.setItem(row_index, 0, self._report_item(item["name"]))
            self.low_stock_table.setItem(row_index, 1, self._report_item(f"{item['stock_quantity']:.2f}", True))
            self.low_stock_table.setItem(row_index, 2, self._report_item(f"{item['reorder_level']:.2f}", True))

        ledger_limit = int(self.ledger_limit_spin.value()) if hasattr(self, "ledger_limit_spin") else 500
        ledger = self.report_service.stock_ledger_between(start_date=start_date, end_date=end_date, limit=ledger_limit)
        self.ledger_table.setRowCount(len(ledger))
        for row_index, row in enumerate(ledger):
            self.ledger_table.setItem(row_index, 0, self._report_item(row["moved_at"]))
            self.ledger_table.setItem(row_index, 1, self._report_item(row["item_name"]))
            self.ledger_table.setItem(row_index, 2, self._report_item(f"{row['quantity_delta']:.2f}", True))
            self.ledger_table.setItem(row_index, 3, self._report_item(row["movement_type"]))
            self.ledger_table.setItem(
                row_index,
                4,
                self._report_item(str(row["reference_id"]) if row["reference_id"] else "-"),
            )
            self.ledger_table.setItem(row_index, 5, self._report_item(row["notes"] or ""))

        trend = self.report_service.sales_trend_between(start_date=start_date, end_date=end_date)
        self.sales_trend_table.setRowCount(len(trend))
        for row_index, row in enumerate(trend):
            self.sales_trend_table.setItem(row_index, 0, self._report_item(row["sale_date"]))
            self.sales_trend_table.setItem(row_index, 1, self._report_item(str(row["bill_count"]), True))
            self.sales_trend_table.setItem(
                row_index,
                2,
                self._report_item(f"{float(row['sales_total']):.2f}", True),
            )
            self.sales_trend_table.setItem(
                row_index,
                3,
                self._report_item(f"{float(row['cogs_total']):.2f}", True),
            )
            self.sales_trend_table.setItem(
                row_index,
                4,
                self._report_item(f"{float(row['gross_profit']):.2f}", True),
            )

        top_limit = int(self.top_items_limit_spin.value()) if hasattr(self, "top_items_limit_spin") else 20
        top_items = self.report_service.top_items_between(start_date=start_date, end_date=end_date, limit=top_limit)
        self.top_items_table.setRowCount(len(top_items))
        for row_index, row in enumerate(top_items):
            self.top_items_table.setItem(row_index, 0, self._report_item(row["name"]))
            self.top_items_table.setItem(
                row_index,
                1,
                self._report_item(f"{float(row['qty_sold']):.2f}", True),
            )
            self.top_items_table.setItem(
                row_index,
                2,
                self._report_item(f"{float(row['sales_value']):.2f}", True),
            )

        if hasattr(self, "audit_log_table"):
            logs = self.bookkeeping_service.list_audit_logs(limit=500)
            self.audit_log_table.setRowCount(len(logs))
            for row_index, log in enumerate(logs):
                self.audit_log_table.setItem(row_index, 0, self._report_item(log.get("created_at", "")))
                self.audit_log_table.setItem(row_index, 1, self._report_item(log.get("actor_role", "")))
                self.audit_log_table.setItem(row_index, 2, self._report_item(log.get("action_type", "")))
                self.audit_log_table.setItem(row_index, 3, self._report_item(log.get("entity_type", "")))
                self.audit_log_table.setItem(row_index, 4, self._report_item(log.get("entity_id", "")))
                self.audit_log_table.setItem(row_index, 5, self._report_item(log.get("details", "")))

        if hasattr(self, "costing_exceptions_table"):
            exceptions = self.report_service.costing_exceptions(limit=400)
            self.costing_exceptions_table.setRowCount(len(exceptions))
            for row_index, ex in enumerate(exceptions):
                self.costing_exceptions_table.setItem(row_index, 0, self._report_item(ex.get("created_at", "")))
                self.costing_exceptions_table.setItem(row_index, 1, self._report_item(ex.get("exception_type", "")))
                self.costing_exceptions_table.setItem(
                    row_index,
                    2,
                    self._report_item(ex.get("item_name") or "-"),
                )
                self.costing_exceptions_table.setItem(
                    row_index,
                    3,
                    self._report_item(str(ex.get("sale_id")) if ex.get("sale_id") else "-"),
                )
                self.costing_exceptions_table.setItem(
                    row_index,
                    4,
                    self._report_item(str(ex.get("item_id")) if ex.get("item_id") else "-"),
                )
                self.costing_exceptions_table.setItem(row_index, 5, self._report_item(ex.get("details", "")))

    def close_day(self) -> None:
        selected_date, ok = QInputDialog.getText(
            self,
            "Close Day",
            "Enter date to close (YYYY-MM-DD):",
            text=date.today().isoformat(),
        )
        if not ok:
            return

        try:
            result = self.bookkeeping_service.close_day(selected_date.strip())
        except ValueError as exc:
            QMessageBox.warning(self, "Close Day", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Close Day Error", str(exc))
            return

        QMessageBox.information(
            self,
            "Day Closed",
            (
                f"Date: {result['closure_date']}\n"
                f"Sales: INR {result['sales']:.2f}\n"
                f"COGS: INR {result['cogs']:.2f}\n"
                f"Expenses: INR {result['expenses']:.2f}\n"
                f"Gross: INR {result['gross_profit']:.2f}\n"
                f"Net: INR {result['net_profit']:.2f}"
            ),
        )
        self._log_audit("day_close", "closure", result["closure_date"], "manual close-day")

    def backup_now(self) -> None:
        try:
            backup_path = create_backup(db_path=self.db_path)
        except Exception as exc:
            QMessageBox.critical(self, "Backup Failed", str(exc))
            return

        self._log_audit("backup_now", "backup", str(backup_path), "manual backup")
        QMessageBox.information(self, "Backup Created", f"Backup saved at:\n{backup_path}")

    def export_backup_dialog(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Select Export Folder")
        if not folder:
            return

        try:
            path = export_backup(destination_dir=folder, db_path=self.db_path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))
            return

        self._log_audit("backup_export", "backup", str(path), f"export folder={folder}")
        QMessageBox.information(self, "Export Complete", f"Backup exported to:\n{path}")

    def restore_backup_dialog(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Backup File", filter="DB Files (*.db)")
        if not file_path:
            return

        pin = self._require_admin_access("Restore Backup")
        if pin is None:
            return

        try:
            current_counts = self.bookkeeping_service.current_database_counts()
            backup_counts = inspect_backup_counts(file_path)
            preview = self._format_restore_preview(
                current_counts=current_counts,
                backup_counts=backup_counts,
                backup_file=file_path,
            )
        except Exception as exc:
            QMessageBox.critical(self, "Restore Pre-check Failed", str(exc))
            return

        confirm = QMessageBox.question(
            self,
            "Confirm Restore",
            preview,
            QMessageBox.Yes | QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        try:
            restore_backup(backup_file=file_path, db_path=self.db_path)
        except Exception as exc:
            QMessageBox.critical(self, "Restore Failed", str(exc))
            return

        QMessageBox.information(
            self,
            "Restore Complete",
            "Backup restored. Please restart the application to reload all data safely.",
        )
        self._log_audit("backup_restore", "backup", file_path, "restore completed")

    def _save_pending_cart(self) -> None:
        self.cart_file.parent.mkdir(parents=True, exist_ok=True)
        if not self.cart:
            self.cart_file.unlink(missing_ok=True)
            return

        payload = {"items": list(self.cart.values())}
        self.cart_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def _load_pending_cart(self) -> None:
        if not self.cart_file.exists():
            return

        try:
            payload = json.loads(self.cart_file.read_text(encoding="utf-8"))
            items = payload.get("items", [])
        except Exception:
            self.cart_file.unlink(missing_ok=True)
            return

        if not items:
            self.cart_file.unlink(missing_ok=True)
            return

        answer = QMessageBox.question(
            self,
            "Recover Pending Bill",
            "Found an unsaved cart from previous session. Restore it?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            self.cart_file.unlink(missing_ok=True)
            return

        self.cart = {int(item["item_id"]): item for item in items}
        self.refresh_cart_table()
